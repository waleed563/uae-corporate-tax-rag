"""
ring1_rag.py — Ring 1 of the improved UAE Corporate Tax RAG system.

WHAT CHANGED FROM baseline_rag.py (read these before you run anything):
─────────────────────────────────────────────────────────────────────────
1. LOADER   : PyMuPDF raw text  →  LangChain PyMuPDFLoader
              Why: preserves page numbers as metadata on every chunk.
              Each chunk now knows which page it came from. Citations
              become "page 20" not "chunk 147".

2. CHUNKER  : Fixed-size 1000 chars  →  RecursiveCharacterTextSplitter
              with smaller chunks (600 chars) and metadata-aware splits.
              Why: 1000-char fixed chunks sliced through Example blocks
              and split a rule from its conditions. 600-char chunks with
              paragraph-first splitting stay inside one idea more often.
              Chunk count will rise (~500 vs 356). That is expected.

3. RETRIEVER: Top-4 cosine only  →  Top-20 cosine then rerank to Top-5
              Why: the right chunk for Q7, Q8, Q15 existed in Qdrant but
              ranked 8th or 9th. Cosine missed it. The reranker reads the
              question AND the chunk together and rescores — it promotes
              the right chunk even when the words don't match well.
              We use a lightweight cross-encoder from sentence-transformers
              (runs locally, no API cost).

4. VECTOR STORE: Tiny numpy store  →  Qdrant in-memory mode
              Why: same speed for one PDF, but Qdrant stores metadata
              (page number, source) with each vector. We use this for
              citations. Also sets us up to swap to a persistent Qdrant
              server later with one line change.

5. PROMPT   : Loose "say I don't know"  →  Stricter 3-rule prompt
              Rule A: explain conditions when answer is conditional —
                      fixes Q13 (had right text, still said "I don't know")
              Rule B: never invent AED amounts not in the context —
                      kills the Q20 hallucination
              Rule C: cite the section name from metadata, not chunk number

WHAT DID NOT CHANGE:
─────────────────────────────────────────────────────────────────────────
- Embedding model: text-embedding-3-small (same — keeps comparison fair)
- LLM: gpt-4o-mini (same)
- Eval questions: same 20 from the spreadsheet
- Output format: same baseline_results.md / .json structure so you can
  compare side by side

LEARNING EXERCISES (do these before running):
─────────────────────────────────────────────────────────────────────────
1. PREDICT THEN RUN
   Read the five CHANGED sections above. Before running, write down:
   - Will chunk count go up or down vs 356? By how much roughly?
   - Which questions do you expect to improve? Which might still fail?
   - What will the reranker do that cosine alone could not?

2. TRACE ONE ITEM
   Pick Q3 ("How much profit before I owe any corporate tax?").
   Follow it through every function manually:
   - load_and_chunk() → what does its chunk look like with metadata?
   - embed_texts() → same function, nothing changed here
   - retrieve_and_rerank() → which 20 come back from Qdrant? which 5
     survive reranking? Does the AED 375,000 chunk make it through?
   - answer_question() → does the new prompt handle it differently?

3. READ THE DIFF
   Open baseline_rag.py and ring1_rag.py side by side in VSCode
   (right-click a file → "Select for Compare", then the other →
   "Compare with Selected"). Read every red/green line. For each
   change ask: why did we make this change?

Run:
    pip install -r requirements_ring1.txt
    python ring1_rag.py --pdf "data/CT General Guide - EN - 10 09 2023.pdf" \
                        --eval uae_tax_rag_eval.xlsx

Outputs:
    ring1_results.md
    ring1_results.json
"""

import argparse
import json
import os
import sys

import numpy as np
from openai import OpenAI
from openpyxl import load_workbook

# CHANGED: LangChain document loader instead of raw PyMuPDF
from langchain_community.document_loaders import PyMuPDFLoader

# CHANGED: same splitter class but different settings
from langchain_text_splitters import RecursiveCharacterTextSplitter

# CHANGED: Qdrant in-memory client instead of tiny numpy store
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct

# CHANGED: cross-encoder reranker (runs locally, no API cost)
from sentence_transformers import CrossEncoder

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
CHUNK_SIZE    = 600        # CHANGED: smaller than baseline (1000) — stays inside one idea
CHUNK_OVERLAP = 100        # CHANGED: tighter overlap
TOP_K_COSINE  = 20         # CHANGED: cast wide net first
TOP_K_RERANK  = 5          # CHANGED: reranker trims to best 5
EMBED_MODEL   = "text-embedding-3-small"   # same as baseline — fair comparison
CHAT_MODEL    = "gpt-4o-mini"              # same as baseline
COLLECTION    = "uae_tax"

# CHANGED: cross-encoder model (downloads ~90MB on first run, cached after)
RERANKER_MODEL = "cross-encoder/ms-marco-MiniLM-L-6-v2"

client = OpenAI()


# ─────────────────────────────────────────────────────────────────────────────
# 1. Load  (CHANGED: LangChain loader preserves page metadata)
# ─────────────────────────────────────────────────────────────────────────────
def load_and_chunk(pdf_path: str) -> list[dict]:
    """
    Returns a list of dicts:
      {"text": "...", "page": 12, "source": "CT General Guide...pdf"}

    Baseline returned a flat list of strings with no metadata.
    Now every chunk knows which page it came from.
    """
    # LangChain loader — each element is a LangChain Document with
    # .page_content (text) and .metadata (page number, source path)
    loader = PyMuPDFLoader(pdf_path)
    pages  = loader.load()   # one Document per page

    # CHANGED: 600-char chunks, paragraph-first splitting
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    docs = splitter.split_documents(pages)  # splits across pages, keeps metadata

    chunks = []
    for doc in docs:
        text = doc.page_content.strip()
        if not text:
            continue
        chunks.append({
            "text":   text,
            "page":   doc.metadata.get("page", 0) + 1,   # 0-indexed → 1-indexed
            "source": os.path.basename(pdf_path),
        })
    return chunks


# ─────────────────────────────────────────────────────────────────────────────
# 2. Embed  (unchanged — same model keeps comparison fair)
# ─────────────────────────────────────────────────────────────────────────────
def embed_texts(texts: list[str]) -> np.ndarray:
    vectors = []
    BATCH = 100
    for i in range(0, len(texts), BATCH):
        batch = texts[i : i + BATCH]
        resp  = client.embeddings.create(model=EMBED_MODEL, input=batch)
        vectors.extend(d.embedding for d in resp.data)
    arr   = np.array(vectors, dtype=np.float32)
    norms = np.linalg.norm(arr, axis=1, keepdims=True)
    norms[norms == 0] = 1e-9
    return arr / norms


# ─────────────────────────────────────────────────────────────────────────────
# 3. Vector store  (CHANGED: Qdrant in-memory instead of tiny numpy class)
# ─────────────────────────────────────────────────────────────────────────────
def build_qdrant(chunks: list[dict], embeddings: np.ndarray) -> QdrantClient:
    """
    Stores chunk text + page metadata in an in-memory Qdrant collection.
    Swap ':memory:' for a real URL to make it persistent later.
    """
    qdrant = QdrantClient(":memory:")
    dim    = embeddings.shape[1]

    qdrant.create_collection(
        collection_name=COLLECTION,
        vectors_config=VectorParams(size=dim, distance=Distance.COSINE),
    )

    points = [
        PointStruct(
            id      = i,
            vector  = embeddings[i].tolist(),
            payload = {"text": chunks[i]["text"],
                       "page": chunks[i]["page"],
                       "source": chunks[i]["source"]},
        )
        for i in range(len(chunks))
    ]
    qdrant.upsert(collection_name=COLLECTION, points=points)
    return qdrant


# ─────────────────────────────────────────────────────────────────────────────
# 4. Retrieve + rerank  (CHANGED: two-stage instead of top-4 cosine only)
# ─────────────────────────────────────────────────────────────────────────────
def retrieve_and_rerank(
    qdrant:   QdrantClient,
    reranker: CrossEncoder,
    question: str,
) -> list[dict]:
    """
    Stage 1 — cosine similarity: retrieve TOP_K_COSINE=20 candidates.
    Stage 2 — cross-encoder rerank: score each (question, chunk) pair
              and keep TOP_K_RERANK=5.

    Why two stages?
    Cosine similarity is fast but compares question and chunk separately.
    The cross-encoder reads them TOGETHER — it understands context better
    but is too slow to run on all 500+ chunks. So: fast cosine narrows
    to 20, then the smarter reranker picks the best 5.
    """
    # Stage 1
    from qdrant_client.models import Query
    q_vec      = embed_texts([question])[0]
    response   = qdrant.query_points(
        collection_name=COLLECTION,
        query=q_vec.tolist(),
        limit=TOP_K_COSINE,
        with_payload=True,
    )
    results = response.points

    # Stage 2
    pairs  = [(question, r.payload["text"]) for r in results]
    scores = reranker.predict(pairs)

    ranked = sorted(
        zip(results, scores),
        key=lambda x: x[1],
        reverse=True,
    )[:TOP_K_RERANK]

    return [
        {
            "rank":         i + 1,
            "chunk_id":     hit.id,
            "cosine_score": round(hit.score, 4),
            "rerank_score": round(float(score), 4),
            "page":         hit.payload["page"],
            "text":         hit.payload["text"],
        }
        for i, (hit, score) in enumerate(ranked)
    ]


# ─────────────────────────────────────────────────────────────────────────────
# 5. Generate  (CHANGED: stricter 3-rule prompt fixes Q13, Q20 failures)
# ─────────────────────────────────────────────────────────────────────────────
GROUNDING_PROMPT = """You are a UAE Corporate Tax assistant. Answer using ONLY the context below.

Rules:
1. If the answer depends on conditions, explain those conditions clearly
   using only the context — do NOT say "I don't know" just because the
   answer is not a simple yes/no.
2. Never state a specific AED amount, penalty figure, percentage, or date
   unless it appears word-for-word in the context below.
3. If the question is genuinely not covered by the context, say:
   "This is not covered in the provided document."
4. End every answer with: Source: page [N] of CTGGCT1.

Context:
{context}

Question: {question}
Answer:"""


def answer_question(
    qdrant:   QdrantClient,
    reranker: CrossEncoder,
    question: str,
) -> dict:
    retrieved = retrieve_and_rerank(qdrant, reranker, question)

    # Build context block — now includes page numbers
    context_blocks = []
    for r in retrieved:
        context_blocks.append(f"[page {r['page']}] {r['text']}")
    context = "\n\n".join(context_blocks)

    resp = client.chat.completions.create(
        model=CHAT_MODEL,
        temperature=0,
        messages=[{
            "role": "user",
            "content": GROUNDING_PROMPT.format(
                context=context,
                question=question,
            ),
        }],
    )
    answer = resp.choices[0].message.content.strip()
    return {"answer": answer, "retrieved": retrieved}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers  (unchanged)
# ─────────────────────────────────────────────────────────────────────────────
def load_questions(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Test Set"]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, question, category = row[0], row[1], row[2]
        if question:
            questions.append({"num": num, "question": question, "category": category})
    wb.close()
    return questions


def write_markdown(results: list[dict], path: str):
    lines = ["# Ring 1 RAG — results to grade", ""]
    lines.append(
        f"Config: chunk_size={CHUNK_SIZE}, overlap={CHUNK_OVERLAP}, "
        f"cosine_k={TOP_K_COSINE}, rerank_k={TOP_K_RERANK}, "
        f"embed={EMBED_MODEL}, chat={CHAT_MODEL}, reranker={RERANKER_MODEL}"
    )
    lines.append("")
    lines.append(
        "For each answer check: did the RIGHT page appear in retrieved chunks? "
        "If yes but answer is wrong → GENERATION failure. "
        "If right page missing → RETRIEVAL failure."
    )
    lines.append("")
    for r in results:
        lines.append(f"## Q{r['num']} — {r['category']}")
        lines.append(f"**Question:** {r['question']}")
        lines.append("")
        lines.append(f"**Answer:** {r['answer']}")
        lines.append("")
        lines.append("**Retrieved chunks (after reranking):**")
        for c in r["retrieved"]:
            preview = c["text"].replace("\n", " ")
            if len(preview) > 300:
                preview = preview[:300] + "…"
            lines.append(
                f"- [rank {c['rank']}] page {c['page']} "
                f"(cosine {c['cosine_score']} → rerank {c['rerank_score']}) {preview}"
            )
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf",  required=True)
    ap.add_argument("--eval", required=True)
    args = ap.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        sys.exit("Set OPENAI_API_KEY first.")

    print("Loading and chunking PDF (LangChain loader)...")
    chunks = load_and_chunk(args.pdf)
    print(f"  {len(chunks)} chunks  (baseline was 356 — more chunks = smaller, focused pieces)")

    print("Embedding chunks...")
    texts      = [c["text"] for c in chunks]
    embeddings = embed_texts(texts)

    print("Building Qdrant in-memory store...")
    qdrant = build_qdrant(chunks, embeddings)

    print(f"Loading cross-encoder reranker ({RERANKER_MODEL})...")
    print("  (downloads ~90MB on first run — cached after)")
    reranker = CrossEncoder(RERANKER_MODEL)

    print("Loading questions...")
    questions = load_questions(args.eval)
    print(f"  {len(questions)} questions")

    results = []
    for q in questions:
        print(f"  Q{q['num']}: {q['question'][:60]}...")
        out = answer_question(qdrant, reranker, q["question"])
        results.append({**q, **out})

    write_markdown(results, "ring1_results.md")
    with open("ring1_results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print("\nDone.")
    print("Open ring1_results.md and grade each answer.")
    print("Compare page citations to baseline chunk citations — that's your story.")
    print("Tag each failure as RETRIEVAL or GENERATION (same as before).")
    print("Then bring both result files back and we read them together.")


if __name__ == "__main__":
    main()