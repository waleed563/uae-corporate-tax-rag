"""
ring3_rag.py — Ring 3 of the improved UAE Corporate Tax RAG system.

WHAT CHANGED FROM ring2_rag.py (read BEFORE running):
─────────────────────────────────────────────────────────────────────────
1. PARENT DOCUMENT CHUNKING  ← the big architectural change
   Ring 2: one level of chunks (600 chars) used for BOTH embedding AND LLM
   Ring 3: two levels of chunks

   CHILD chunks (300 chars) — small, focused, used for EMBEDDING + SEARCH
   PARENT chunks (~2,000 chars) — full paragraphs, used for LLM CONTEXT

   Why: document analysis revealed average paragraph is 2,146 chars.
   Ring 2's 600-char chunks were cutting through 95% of paragraphs.
   The LLM was reading fragments, not complete ideas.

   How retrieval works now:
     Search children (tight match) → look up their parent (full context)
     → rerank parents → compress parents → LLM reads complete paragraphs

2. EXAMPLE-BLOCK PRESERVATION
   Ring 2: "Example N:" blocks split across 3-4 chunks
   Ring 3: detected with regex, kept as ONE parent chunk regardless of size

   Why: 36 worked Examples in this document each have a setup paragraph
   and a calculation/result. Splitting them apart means the LLM sees
   the setup without the result or vice versa. Q14 failure was partly
   caused by this.

3. COMPRESSOR UPDATED FOR PARENT SIZE
   Ring 2: compressed 600-char chunks (sometimes stripped too much)
   Ring 3: compresses 2,000-char parents (more content = smarter compression)

   Threshold lowered 0.35 → 0.30: parents have more sentences to judge from,
   so a lower threshold is safe — we're not starving the model of context.
   MIN_SENTENCES_KEPT raised 2 → 3: parents have ~8-12 sentences,
   keeping at least 3 preserves enough context while removing noise.

WHAT DID NOT CHANGE FROM RING 2:
─────────────────────────────────────────────────────────────────────────
- Multi-query (3 variants per question) — same
- Embedding model: text-embedding-3-small — same
- Qdrant in-memory — same (children stored, parents in a Python dict)
- Reranker: cross-encoder/ms-marco-MiniLM-L-6-v2 — same
- LangSmith observability — same (env var setup)
- Token budget — same (3,500 tokens, slightly higher for richer context)
- LLM + guardrails prompt — same (rules A, B, C unchanged)

LEARNING EXERCISES (do before running):
─────────────────────────────────────────────────────────────────────────
1. PREDICT THEN RUN
   Parent chunks are ~2,000 chars (full paragraphs). Children are 300 chars.
   The system finds small children but sends large parents to the LLM.
   Predict:
   - Will chunk count go up or down compared to Ring 2 (which had ~500)?
     Think: MORE children (smaller) but from FEWER parents (larger paragraphs).
   - Q14 (free-zone selling to mainland): the condition and its Example were
     in separate chunks in Ring 2. Will parent chunking fix that?
   - Will token count per question go UP or DOWN after compression?
     Parents are larger but compression is more aggressive (0.30 threshold).

2. TRACE ONE ITEM
   After running, open LangSmith and find Q5 (free-zone 2M profit).
   This is a conditional question that spans a rule paragraph AND an Example.
   - Did retrieval find an Example-type parent?
   - How many sentences survived compression?
   - Is the answer more complete than Ring 2's version?

3. READ THE DIFF
   Open ring2_rag.py and ring3_rag.py side by side in VSCode.
   Find these functions and understand what changed:
   - build_parent_child_chunks()  ← NEW: replaces load_and_chunk()
   - build_qdrant()               ← CHANGED: stores children, not parents
   - retrieve_and_rerank()        ← CHANGED: child→parent lookup step added
   - compress_chunks()            ← CHANGED: lower threshold, higher minimum

Run:
    python ring3_rag.py --pdf "data/CT General Guide - EN - 10 09 2023.pdf" \\
                        --eval uae_tax_rag_eval.xlsx

Outputs:
    ring3_results.md
    ring3_results.json
"""

import argparse
import json
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor

import numpy as np
import tiktoken
from openai import OpenAI
from openpyxl import load_workbook
from langchain_community.document_loaders import PyMuPDFLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
from sentence_transformers import CrossEncoder

# FIX 3: pysbd for robust sentence splitting — handles "AED 375,000." and footnotes
# correctly unlike naive split(". ") which creates false empty sentences.
try:
    from pysbd import Segmenter as PySBDSegmenter
    _segmenter = PySBDSegmenter(language="en", clean=True)
    def split_sentences(text: str) -> list[str]:
        return [s.strip() for s in _segmenter.segment(text) if s.strip() and len(s.strip()) > 8]
    PYSBD_AVAILABLE = True
except ImportError:
    PYSBD_AVAILABLE = False
    def split_sentences(text: str) -> list[str]:
        # Fallback: regex-based split that handles common legal text patterns
        import re
        raw = re.split(r'(?<=[.!?])\s+(?=[A-Z0-9])', text)
        return [s.strip() for s in raw if s.strip() and len(s.strip()) > 8]

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────────────────────────────────────────
# LangSmith (same as Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
LANGSMITH_ENABLED = bool(os.getenv("LANGCHAIN_API_KEY"))
if LANGSMITH_ENABLED:
    os.environ.setdefault("LANGCHAIN_TRACING_V2", "true")
    os.environ.setdefault("LANGCHAIN_PROJECT", "uae-tax-rag")
    print("LangSmith tracing: ON")
else:
    print("LangSmith tracing: OFF — add LANGCHAIN_API_KEY to .env to enable")

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
PARENT_CHUNK_SIZE    = 2000   # CHANGED: full paragraph size (avg para = 2,146 chars)
PARENT_OVERLAP       = 200    # CHANGED: 10% overlap between parents
CHILD_CHUNK_SIZE     = 300    # CHANGED: small chunks for tight embeddings
CHILD_OVERLAP        = 50     # CHANGED: small overlap for children
RERANKER_MAX_CHARS   = 1000   # CHANGED: truncate parent for reranker input (model sweet spot)
TOP_K_COSINE                = 30    # FIX 4: 20→30 — wider net catches Q7/Q8/Q20 regressions
TOP_K_RERANK                = 5     # same as Ring 2
NUM_QUERY_VARIANTS          = 3     # same as Ring 2
COMPRESS_THRESHOLD_PROSE    = 0.28  # FIX 2a: type-aware — prose chunks are focused, be lenient
COMPRESS_THRESHOLD_EXAMPLE  = 0.40  # FIX 2b: type-aware — Examples have 100+ sentences, be strict
MAX_SENTENCES_PER_CHUNK     = 20    # FIX 2c: hard cap — no single chunk can blow the token budget
MIN_SENTENCES_KEPT          = 3     # same — never compress below 3 sentences
TOKEN_BUDGET                = 3500  # same
EMBED_MODEL          = "text-embedding-3-small"
CHAT_MODEL           = "gpt-4o-mini"
COLLECTION           = "uae_tax_ring3"
RERANKER_MODEL       = "cross-encoder/ms-marco-MiniLM-L-6-v2"

client    = OpenAI()
try:
    tokenizer = tiktoken.encoding_for_model("gpt-4o-mini")
except Exception:
    tokenizer = None

# FIX 5: safe token counter — tiktoken can fail on restricted networks;
# fall back to char/4 estimate which is accurate enough for budget decisions.
def count_tokens(text: str) -> int:
    if tokenizer is not None:
        try:
            return len(tokenizer.encode(text))
        except Exception:
            pass
    return max(1, len(text) // 4)   # ~4 chars per token for English text


# ─────────────────────────────────────────────────────────────────────────────
# 1. Build parent + child chunks  (CHANGED: replaces load_and_chunk)
# ─────────────────────────────────────────────────────────────────────────────
def build_parent_child_chunks(pdf_path: str) -> tuple[dict, list[dict]]:
    """
    Two-level chunking strategy informed by document analysis:

    PARENTS (~2,000 chars each):
    - Detected "Example N:" blocks → kept as ONE chunk regardless of size
      (36 Examples in this document; each has setup + calculation + result)
    - All other text → split at paragraph/sentence boundaries into ~2,000 chars
      (avg paragraph in this doc is 2,146 chars, so one parent ≈ one paragraph)

    CHILDREN (300 chars each):
    - Every parent is split into small 300-char children
    - Children are what gets embedded and searched in Qdrant
    - Each child stores its parent_id so we can look up the full parent

    Returns:
        parent_store: {parent_id → {text, page, type, source}}
        children:     [{id, parent_id, text, page}]
    """
    # Load pages (same loader as Ring 2)
    loader = PyMuPDFLoader(pdf_path)
    pages  = loader.load()

    # Build full text + track where each page ends (for page estimation)
    page_texts     = [p.page_content for p in pages]
    page_nums      = [p.metadata.get("page", i) + 1 for i, p in enumerate(pages)]
    cumulative     = 0
    page_boundaries = []
    for text, num in zip(page_texts, page_nums):
        cumulative += len(text) + 1
        page_boundaries.append((cumulative, num))

    full_text = "\n".join(page_texts)

    def estimate_page(char_pos: int) -> int:
        for boundary, page_num in page_boundaries:
            if char_pos <= boundary:
                return page_num
        return page_boundaries[-1][1]

    # ── Split: detect Example blocks vs prose ──────────────────────────────
    # re.split with capturing group → alternates [prose, "Example N:", content, ...]
    parts = re.split(r"(Example \d+:)", full_text)

    prose_splitter = RecursiveCharacterTextSplitter(
        chunk_size=PARENT_CHUNK_SIZE,
        chunk_overlap=PARENT_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    child_splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHILD_CHUNK_SIZE,
        chunk_overlap=CHILD_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )

    raw_parents = []

    i = 0
    while i < len(parts):
        part = parts[i]

        if re.match(r"Example \d+:", part):
            # Combine "Example N:" header with its content (next part)
            content = parts[i + 1] if i + 1 < len(parts) else ""
            full_example = (part + content).strip()
            if full_example:
                pos = full_text.find(part)
                raw_parents.append({
                    "text": full_example,
                    "page": estimate_page(pos) if pos >= 0 else 1,
                    "type": "example",     # preserved whole
                })
            i += 2   # skip the content part too
        else:
            # Regular prose → split into parent-sized chunks
            if part.strip():
                for prose_parent in prose_splitter.split_text(part):
                    if prose_parent.strip():
                        pos = full_text.find(prose_parent[:80])
                        raw_parents.append({
                            "text": prose_parent.strip(),
                            "page": estimate_page(pos) if pos >= 0 else 1,
                            "type": "prose",
                        })
            i += 1

    # ── Build parent store + child chunks ─────────────────────────────────
    parent_store: dict[str, dict] = {}
    children:     list[dict]      = []

    for idx, parent in enumerate(raw_parents):
        pid = f"p{idx}"
        parent_store[pid] = {
            **parent,
            "source": os.path.basename(pdf_path),
        }
        for j, child_text in enumerate(child_splitter.split_text(parent["text"])):
            if child_text.strip():
                children.append({
                    "id":        f"c{idx}_{j}",
                    "parent_id": pid,
                    "text":      child_text.strip(),
                    "page":      parent["page"],
                })

    return parent_store, children


# ─────────────────────────────────────────────────────────────────────────────
# 2. Embed  (unchanged from Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
def embed_texts(texts: list[str]) -> np.ndarray:
    vectors = []
    BATCH   = 100
    for i in range(0, len(texts), BATCH):
        batch = texts[i : i + BATCH]
        resp  = client.embeddings.create(model=EMBED_MODEL, input=batch)
        vectors.extend(d.embedding for d in resp.data)
    arr   = np.array(vectors, dtype=np.float32)
    norms = np.linalg.norm(arr, axis=1, keepdims=True)
    norms[norms == 0] = 1e-9
    return arr / norms


# ─────────────────────────────────────────────────────────────────────────────
# 3. Qdrant store  (CHANGED: stores CHILDREN, not parents)
# ─────────────────────────────────────────────────────────────────────────────
def build_qdrant(children: list[dict], embeddings: np.ndarray) -> QdrantClient:
    """
    Stores child embeddings. Each child's payload includes parent_id
    so we can look up the full parent after retrieval.
    """
    qdrant = QdrantClient(":memory:")
    dim    = embeddings.shape[1]
    qdrant.create_collection(
        collection_name=COLLECTION,
        vectors_config=VectorParams(size=dim, distance=Distance.COSINE),
    )
    points = [
        PointStruct(
            id      = i,
            vector  = embeddings[i].tolist(),
            payload = {
                "text":      children[i]["text"],
                "parent_id": children[i]["parent_id"],
                "page":      children[i]["page"],
            },
        )
        for i in range(len(children))
    ]
    qdrant.upsert(collection_name=COLLECTION, points=points)
    return qdrant


# ─────────────────────────────────────────────────────────────────────────────
# 4a. Multi-query — generate alternate phrasings (same as Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
QUERY_EXPANSION_PROMPT = """You are helping retrieve information from a UAE Corporate Tax document.

Given the user's question below, generate {n} alternate phrasings of the same question
using formal legal and tax terminology that might appear in a regulatory document.
Return ONLY the alternate questions, one per line. No numbering, no explanations.

User question: {question}"""


def generate_query_variants(question: str) -> list[str]:
    resp = client.chat.completions.create(
        model=CHAT_MODEL,
        temperature=0.3,
        messages=[{
            "role": "user",
            "content": QUERY_EXPANSION_PROMPT.format(
                n=NUM_QUERY_VARIANTS,
                question=question,
            ),
        }],
    )
    raw      = resp.choices[0].message.content.strip()
    variants = [line.strip() for line in raw.split("\n") if line.strip()]
    return variants[:NUM_QUERY_VARIANTS]


# ─────────────────────────────────────────────────────────────────────────────
# 4b. Retrieve children for one query
# ─────────────────────────────────────────────────────────────────────────────
def retrieve_one(qdrant: QdrantClient, query: str) -> list:
    q_vec    = embed_texts([query])[0]
    response = qdrant.query_points(
        collection_name=COLLECTION,
        query=q_vec.tolist(),
        limit=TOP_K_COSINE,
        with_payload=True,
    )
    return response.points


# ─────────────────────────────────────────────────────────────────────────────
# 4c. Multi-query retrieve → parent lookup → rerank  (CHANGED: parent lookup added)
# ─────────────────────────────────────────────────────────────────────────────
def retrieve_and_rerank(
    qdrant:       QdrantClient,
    reranker:     CrossEncoder,
    question:     str,
    variants:     list[str],
    parent_store: dict,
) -> list[dict]:
    """
    Step 1 — Parallel child retrieval across original + variant queries.
    Step 2 — Parent lookup: for each child hit, fetch its parent chunk.
              Multiple children can point to the same parent — deduplicate.
    Step 3 — Rerank (question, parent_text) pairs using cross-encoder.
              Parent is truncated to RERANKER_MAX_CHARS (cross-encoder sweet spot).
    Step 4 — Return top TOP_K_RERANK parents with full text for the LLM.

    Why parent lookup matters:
    If a child chunk (300 chars) matches the query, its parent (2,000 chars)
    contains the FULL paragraph context. The LLM gets the complete idea,
    not just the fragment that happened to match the embedding.
    """
    all_queries = [question] + variants

    # Step 1 — parallel child retrieval
    with ThreadPoolExecutor(max_workers=len(all_queries)) as pool:
        futures            = [pool.submit(retrieve_one, qdrant, q) for q in all_queries]
        results_per_query  = [f.result() for f in futures]

    # Step 2 — parent lookup + deduplication
    seen_parents   = set()
    unique_parents = []
    for results in results_per_query:
        for hit in results:
            pid = hit.payload["parent_id"]
            if pid not in seen_parents:
                seen_parents.add(pid)
                parent = parent_store.get(pid)
                if parent:
                    unique_parents.append({
                        "parent_id":    pid,
                        "text":         parent["text"],
                        "page":         parent["page"],
                        "type":         parent.get("type", "prose"),
                        "cosine_score": round(hit.score, 4),
                    })

    if not unique_parents:
        return []

    # Step 3 — rerank against ORIGINAL question (not variants)
    # Truncate parent for reranker — cross-encoder trained on short passages
    pairs  = [(question, p["text"][:RERANKER_MAX_CHARS]) for p in unique_parents]
    scores = reranker.predict(pairs)

    ranked = sorted(
        zip(unique_parents, scores),
        key=lambda x: x[1],
        reverse=True,
    )[:TOP_K_RERANK]

    return [
        {
            "rank":         i + 1,
            "parent_id":    p["parent_id"],
            "rerank_score": round(float(score), 4),
            "page":         p["page"],
            "type":         p["type"],
            "text":         p["text"],   # full parent text → to LLM
        }
        for i, (p, score) in enumerate(ranked)
    ]


# ─────────────────────────────────────────────────────────────────────────────
# 5. Contextual compression  (FIXED: pysbd splitting, type-aware thresholds, MAX cap)
# ─────────────────────────────────────────────────────────────────────────────
def compress_chunks(question: str, chunks: list[dict]) -> list[dict]:
    """
    FIXES applied vs original Ring 3:

    FIX 2a — Type-aware thresholds:
      prose:   0.28 (more lenient — focused paragraphs, every sentence counts)
      example: 0.40 (more strict — 100+ sentence examples need aggressive filtering)

    FIX 2c — MAX_SENTENCES_PER_CHUNK = 20:
      After filtering by threshold, cap at 20 sentences.
      Prevents one massive Example block from consuming the entire token budget.
      In the first run: Q13/Q14 rank-1 chunks had 85 sentences, exceeding the
      3,500-token budget by themselves → budgeted returned empty → 0 context.

    FIX 3 — pysbd sentence splitting:
      split(". ") incorrectly split "AED 375,000." into ["AED 375,000", ""] and
      created footnote markers like "1 Article 57" as fake sentences.
      pysbd handles legal/regulatory text correctly.
    """
    q_vec = embed_texts([question])[0]

    compressed = []
    for chunk in chunks:
        # FIX 3: use pysbd for accurate legal text sentence segmentation
        chunk_type = chunk.get("type", "prose")
        sentences  = split_sentences(chunk["text"])

        if len(sentences) <= 2:
            # Too short to compress meaningfully — keep as-is
            compressed.append({**chunk, "compressed": False})
            continue

        # FIX 2a: use type-aware threshold
        threshold = (
            COMPRESS_THRESHOLD_EXAMPLE if chunk_type == "example"
            else COMPRESS_THRESHOLD_PROSE
        )

        sent_vecs = embed_texts(sentences)
        scored    = [
            (sent, float(np.dot(q_vec, vec)))
            for sent, vec in zip(sentences, sent_vecs)
        ]

        kept = [sent for sent, score in scored if score >= threshold]

        # Floor: never compress below MIN_SENTENCES_KEPT
        if len(kept) < MIN_SENTENCES_KEPT:
            top_sents = {
                s for s, _ in
                sorted(scored, key=lambda x: x[1], reverse=True)[:MIN_SENTENCES_KEPT]
            }
            kept = [sent for sent in sentences if sent in top_sents]

        # FIX 2c: ceiling — cap at MAX_SENTENCES_PER_CHUNK
        # Keep the HIGHEST-scoring sentences so we don't lose the most relevant ones
        if len(kept) > MAX_SENTENCES_PER_CHUNK:
            # Build a score lookup and take the top MAX by score,
            # then reorder by original position for readability
            score_map = {sent: score for sent, score in scored}
            top_kept  = sorted(kept, key=lambda s: score_map.get(s, 0), reverse=True)[:MAX_SENTENCES_PER_CHUNK]
            kept      = [s for s in sentences if s in set(top_kept)]

        if not kept:
            compressed.append({**chunk, "compressed": False})
        else:
            compressed.append({
                **chunk,
                "text":       " ".join(kept),   # join with space (pysbd segments include punctuation)
                "compressed": True,
                "kept_sents": len(kept),
                "orig_sents": len(sentences),
            })

    return compressed


# ─────────────────────────────────────────────────────────────────────────────
# 6. Token budget  (FIXED: safe counter, truncate-not-drop, never empty)
# ─────────────────────────────────────────────────────────────────────────────
def apply_token_budget(chunks: list[dict], budget: int) -> list[dict]:
    """
    FIXES applied vs original Ring 3:

    FIX 1a — Truncate first chunk instead of dropping:
      Old code did `break` when first chunk exceeded budget.
      If rank-1 chunk (often a large Example) was 4,000 tokens, ALL chunks
      were dropped → budgeted returned [] → 0 context → LLM hallucinated.
      Fix: if first chunk alone exceeds budget, truncate its text to fit.

    FIX 1b — Never return empty list:
      After all fixes, if budgeted is still somehow empty (shouldn't happen),
      force-include a truncated version of the first chunk.
      This is a last-resort safety net. LLM getting imperfect context is
      always better than LLM getting NO context.

    FIX 5 — Uses count_tokens() wrapper:
      Handles tiktoken failures gracefully with char-count fallback.
    """
    if not chunks:
        return []

    kept  = []
    total = 0

    for i, chunk in enumerate(chunks):
        tokens = count_tokens(chunk["text"])

        if total + tokens <= budget:
            kept.append(chunk)
            total += tokens
        elif i == 0:
            # FIX 1a: first chunk exceeds budget — truncate it to fit
            # Take as many characters as the budget allows (~4 chars per token)
            max_chars = budget * 4
            truncated = {
                **chunk,
                "text": chunk["text"][:max_chars] + "…",
            }
            kept.append(truncated)
            total = budget
            break
        else:
            # Subsequent chunk doesn't fit — stop here
            break

    # FIX 1b: last-resort safety net
    if not kept and chunks:
        max_chars = budget * 4
        kept = [{**chunks[0], "text": chunks[0]["text"][:max_chars] + "…"}]

    return kept


# ─────────────────────────────────────────────────────────────────────────────
# 7. Generate answer  (prompt same as Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
GROUNDING_PROMPT = """You are a UAE Corporate Tax assistant. Answer using ONLY the context below.

Rules:
1. If the answer depends on conditions, explain those conditions clearly
   using only the context — do NOT say "I don't know" just because the
   answer is not a simple yes/no.
2. Never state a specific AED amount, penalty figure, percentage, or date
   unless it appears word-for-word in the context below.
3. If the question is genuinely not covered by the context, say:
   "This is not covered in the provided document."
4. End every answer with: Source: page [N] of CTGGCT1.

Context:
{context}

Question: {question}
Answer:"""


def answer_question(
    qdrant:       QdrantClient,
    reranker:     CrossEncoder,
    question:     str,
    parent_store: dict,
) -> dict:
    # Step 1 — generate variants (1 LLM call)
    variants = generate_query_variants(question)

    # Step 2 — multi-query → parent lookup → rerank
    retrieved = retrieve_and_rerank(qdrant, reranker, question, variants, parent_store)

    # Step 3 — compress parents
    compressed = compress_chunks(question, retrieved)

    # Step 4 — token budget
    budgeted = apply_token_budget(compressed, TOKEN_BUDGET)

    # Step 5 — build context with page + type tags
    context_blocks = [
        f"[page {c['page']}, {c['type']}] {c['text']}"
        for c in budgeted
    ]
    context = "\n\n".join(context_blocks)

    # FIX 6: explicit guard — if context is empty despite all fixes, surface the problem
    # clearly rather than silently sending blank context to the LLM.
    if not context.strip():
        context = "No relevant context was retrieved for this question."

    token_count = count_tokens(context)   # FIX 5: use safe counter

    # Step 6 — generate answer (1 LLM call)
    resp = client.chat.completions.create(
        model=CHAT_MODEL,
        temperature=0,
        messages=[{
            "role": "user",
            "content": GROUNDING_PROMPT.format(
                context=context,
                question=question,
            ),
        }],
    )
    answer = resp.choices[0].message.content.strip()

    return {
        "answer":      answer,
        "retrieved":   retrieved,
        "compressed":  compressed,
        "variants":    variants,
        "token_count": token_count,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Helpers  (same as Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
def load_questions(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Test Set"]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, question, category = row[0], row[1], row[2]
        if question:
            questions.append({"num": num, "question": question, "category": category})
    wb.close()
    return questions


def write_markdown(results: list[dict], path: str):
    lines = ["# Ring 3 RAG — results to grade", ""]
    lines.append(
        f"Config: parent={PARENT_CHUNK_SIZE}, child={CHILD_CHUNK_SIZE}, "
        f"cosine_k={TOP_K_COSINE}, rerank_k={TOP_K_RERANK}, "
        f"variants={NUM_QUERY_VARIANTS}, "
        f"compress_prose={COMPRESS_THRESHOLD_PROSE}/example={COMPRESS_THRESHOLD_EXAMPLE}, "
        f"max_sents={MAX_SENTENCES_PER_CHUNK}, token_budget={TOKEN_BUDGET}, "
        f"pysbd={PYSBD_AVAILABLE}"
    )
    lines.append("")
    for r in results:
        lines.append(f"## Q{r['num']} — {r['category']}")
        lines.append(f"**Question:** {r['question']}")
        lines.append("")
        lines.append("**Query variants:**")
        for v in r.get("variants", []):
            lines.append(f"- {v}")
        lines.append("")
        lines.append(f"**Tokens in context:** {r.get('token_count', '?')}")
        lines.append("")
        lines.append(f"**Answer:** {r['answer']}")
        lines.append("")
        lines.append("**Retrieved parents (after reranking + compression):**")
        for c in r.get("compressed", []):
            preview = c["text"].replace("\n", " ")
            if len(preview) > 300:
                preview = preview[:300] + "…"
            comp_note = (
                f"compressed {c['orig_sents']}→{c['kept_sents']} sentences"
                if c.get("compressed") else "kept as-is"
            )
            lines.append(
                f"- [rank {c['rank']}] page {c['page']} [{c['type']}] "
                f"(rerank {c['rerank_score']}, {comp_note}) {preview}"
            )
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf",  required=True)
    ap.add_argument("--eval", required=True)
    args = ap.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        sys.exit("Set OPENAI_API_KEY first.")

    print("Building parent + child chunks...")
    parent_store, children = build_parent_child_chunks(args.pdf)
    prose_count   = sum(1 for p in parent_store.values() if p["type"] == "prose")
    example_count = sum(1 for p in parent_store.values() if p["type"] == "example")
    print(f"  {len(parent_store)} parents ({prose_count} prose, {example_count} examples kept whole)")
    print(f"  {len(children)} children  (these get embedded)")

    print("Embedding children...")
    child_texts  = [c["text"] for c in children]
    embeddings   = embed_texts(child_texts)

    print("Building Qdrant store (children only)...")
    qdrant = build_qdrant(children, embeddings)

    print(f"Loading reranker ({RERANKER_MODEL})...")
    reranker = CrossEncoder(RERANKER_MODEL)

    print("Loading questions...")
    questions = load_questions(args.eval)
    print(f"  {len(questions)} questions\n")

    results = []
    for q in questions:
        print(f"  Q{q['num']}: {q['question'][:55]}...")
        out = answer_question(qdrant, reranker, q["question"], parent_store)
        tok = out.get("token_count", "?")
        print(f"         → {tok} tokens in context")
        results.append({**q, **out})

    write_markdown(results, "ring3_results.md")
    with open("ring3_results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print("\nDone.")
    print("Key things to check in ring3_results.md:")
    print("  1. Are Example-type parents appearing for calculation questions?")
    print("  2. Is Q14 (free-zone mainland sales) now pulling the right Example block?")
    print("  3. Token counts — are they higher or lower than Ring 2?")
    print("Then run: python evaluate_tax_v2.py (update RESULT_FILES to include ring3)")


if __name__ == "__main__":
    main()