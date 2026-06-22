"""
baseline_rag.py — Deliberately NAIVE RAG over a single UAE Corporate Tax PDF (CTGGCT1).

This is the BASELINE for your eval. The whole point is that it is simple and that
it FAILS in visible ways (fixed-size chunking will split tables and separate a rule
from its conditions). Do NOT make it clever here — measure it, then build the
improved version in a separate file so the before/after is honest.

What it does:
  1. Loads raw text from the PDF
  2. Chunks it with fixed-size splitting (the deliberate weak point)
  3. Embeds the chunks (one model, same for query + chunks)
  4. Stores them in a tiny in-memory cosine-similarity index (no vector DB needed for one PDF)
  5. Retrieves top-k chunks for each question
  6. Generates an answer that is told to ONLY use the context and to say "I don't know" otherwise

It logs the RETRIEVED CHUNKS for every question, so when an answer is wrong you can tell
whether it was a RETRIEVAL failure (right text never pulled) or a GENERATION failure
(right text pulled, model still messed up). That distinction is your diagnostic story.

Setup:
    pip install -r requirements.txt
    export OPENAI_API_KEY=sk-...        # or set it in your shell

Run:
    python baseline_rag.py --pdf CTGGCT1.pdf --eval uae_tax_rag_eval.xlsx

Outputs (next to the script):
    baseline_results.md    human-readable: each question, the answer, and the chunks retrieved
    baseline_results.json  machine-readable log for later analysis / charts
"""

import argparse
import json
import os
import sys

import numpy as np
import fitz  # PyMuPDF
from openai import OpenAI
from openpyxl import load_workbook
from langchain_text_splitters import RecursiveCharacterTextSplitter

# Optional: load OPENAI_API_KEY from a local .env file if python-dotenv is installed
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Config — the knobs you'll later change in the IMPROVED version
# ---------------------------------------------------------------------------
CHUNK_SIZE = 1000          # characters; fixed-size = deliberate baseline weakness
CHUNK_OVERLAP = 150
TOP_K = 4                  # how many chunks to retrieve per question
EMBED_MODEL = "text-embedding-3-small"
CHAT_MODEL = "gpt-4o-mini"

client = OpenAI()  # reads OPENAI_API_KEY from the environment


# ---------------------------------------------------------------------------
# 1. Load
# ---------------------------------------------------------------------------
def load_pdf_text(path: str) -> str:
    doc = fitz.open(path)
    pages = [page.get_text() for page in doc]
    doc.close()
    return "\n".join(pages)


# ---------------------------------------------------------------------------
# 2. Chunk (naive, on purpose)
# ---------------------------------------------------------------------------
def chunk_text(text: str) -> list[str]:
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    return [c.strip() for c in splitter.split_text(text) if c.strip()]


# ---------------------------------------------------------------------------
# 3. Embed
# ---------------------------------------------------------------------------
def embed_texts(texts: list[str]) -> np.ndarray:
    """Embed a list of strings, batched. Returns an (n, d) float32 array."""
    vectors = []
    BATCH = 100
    for i in range(0, len(texts), BATCH):
        batch = texts[i : i + BATCH]
        resp = client.embeddings.create(model=EMBED_MODEL, input=batch)
        vectors.extend(d.embedding for d in resp.data)
    arr = np.array(vectors, dtype=np.float32)
    # normalize so dot product == cosine similarity
    norms = np.linalg.norm(arr, axis=1, keepdims=True)
    norms[norms == 0] = 1e-9
    return arr / norms


# ---------------------------------------------------------------------------
# 4. Tiny in-memory store  (swap for Qdrant in the improved version)
# ---------------------------------------------------------------------------
class TinyVectorStore:
    def __init__(self, chunks: list[str], embeddings: np.ndarray):
        self.chunks = chunks
        self.embeddings = embeddings  # already normalized

    def search(self, query_vec: np.ndarray, k: int) -> list[tuple[int, float]]:
        scores = self.embeddings @ query_vec  # cosine similarity
        top = np.argsort(scores)[::-1][:k]
        return [(int(i), float(scores[i])) for i in top]


# ---------------------------------------------------------------------------
# 5 + 6. Retrieve and generate
# ---------------------------------------------------------------------------
GROUNDING_PROMPT = """You are answering questions about UAE Corporate Tax using ONLY the context below.

Rules:
- Answer ONLY from the context. Do not use outside knowledge.
- If the answer is not in the context, reply exactly: "I don't know based on this document."
- Cite the chunk number(s) you used, like [chunk 2].
- Be concise.

Context:
{context}

Question: {question}
Answer:"""


def answer_question(store: TinyVectorStore, question: str) -> dict:
    q_vec = embed_texts([question])[0]
    hits = store.search(q_vec, TOP_K)

    retrieved = []
    context_blocks = []
    for rank, (idx, score) in enumerate(hits, start=1):
        chunk = store.chunks[idx]
        retrieved.append({"rank": rank, "chunk_id": idx, "score": round(score, 4), "text": chunk})
        context_blocks.append(f"[chunk {rank}] {chunk}")

    context = "\n\n".join(context_blocks)
    resp = client.chat.completions.create(
        model=CHAT_MODEL,
        temperature=0,
        messages=[{"role": "user", "content": GROUNDING_PROMPT.format(context=context, question=question)}],
    )
    answer = resp.choices[0].message.content.strip()
    return {"answer": answer, "retrieved": retrieved}


# ---------------------------------------------------------------------------
# Load the 20 questions from the eval spreadsheet
# ---------------------------------------------------------------------------
def load_questions(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Test Set"]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, question, category = row[0], row[1], row[2]
        if question:
            questions.append({"num": num, "question": question, "category": category})
    wb.close()
    return questions


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
def write_markdown(results: list[dict], path: str):
    lines = ["# Baseline RAG — results to grade", ""]
    lines.append(f"Config: chunk_size={CHUNK_SIZE}, overlap={CHUNK_OVERLAP}, "
                 f"top_k={TOP_K}, embed={EMBED_MODEL}, chat={CHAT_MODEL}")
    lines.append("")
    lines.append("For each failure, check the retrieved chunks: if the correct text is "
                 "NOT below, it's a RETRIEVAL failure. If it IS below but the answer is "
                 "still wrong, it's a GENERATION failure.")
    lines.append("")
    for r in results:
        lines.append(f"## Q{r['num']} — {r['category']}")
        lines.append(f"**Question:** {r['question']}")
        lines.append("")
        lines.append(f"**Answer:** {r['answer']}")
        lines.append("")
        lines.append("**Retrieved chunks:**")
        for c in r["retrieved"]:
            preview = c["text"].replace("\n", " ")
            if len(preview) > 300:
                preview = preview[:300] + "…"
            lines.append(f"- [chunk {c['rank']}] (score {c['score']}) {preview}")
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", required=True, help="Path to the CTGGCT1 PDF")
    ap.add_argument("--eval", required=True, help="Path to uae_tax_rag_eval.xlsx")
    args = ap.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        sys.exit("Set OPENAI_API_KEY first:  export OPENAI_API_KEY=sk-...")

    print("Loading PDF...")
    text = load_pdf_text(args.pdf)
    print(f"  {len(text):,} characters")

    print("Chunking...")
    chunks = chunk_text(text)
    print(f"  {len(chunks)} chunks")

    print("Embedding chunks (one-time)...")
    embeddings = embed_texts(chunks)
    store = TinyVectorStore(chunks, embeddings)

    print("Loading questions...")
    questions = load_questions(args.eval)
    print(f"  {len(questions)} questions")

    results = []
    for q in questions:
        print(f"  Q{q['num']}: {q['question'][:60]}...")
        out = answer_question(store, q["question"])
        results.append({**q, **out})

    write_markdown(results, "baseline_results.md")
    with open("baseline_results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print("\nDone. Read baseline_results.md, grade each answer in the spreadsheet,")
    print("and tag every failure as RETRIEVAL or GENERATION.")


if __name__ == "__main__":
    main()