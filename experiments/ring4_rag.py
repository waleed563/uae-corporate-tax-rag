"""
ring4_rag.py — Ring 4 of the UAE Corporate Tax RAG system.

WHAT CHANGED FROM ring3_rag.py (read BEFORE running):
─────────────────────────────────────────────────────────────────────────
1. HYDE (HYPOTHETICAL DOCUMENT EMBEDDINGS)
   Ring 3: embed(question) → retrieve
   Ring 4: LLM generates a hypothetical answer → embed(hypothesis) → retrieve
            ALSO embed(question) → retrieve → union both result sets before reranking

   Why: The UAE CT document uses legal phrasing ("natural person conducting
   business without commercial registration") but users ask in plain English
   ("I freelance from home"). A hypothetical answer written in document-style
   language bridges this vocabulary gap far better than multi-query alone.
   The question embedding runs in parallel so we don't lose the direct-match
   signal either.

2. UPGRADED EMBEDDING MODEL: text-embedding-3-large
   Ring 3: text-embedding-3-small (1536 dims)
   Ring 4: text-embedding-3-large (3072 dims)

   Why: Richer semantic representations, especially for legal/tax terminology.
   Direct improvement to cosine similarity quality.
   Cost: ~6× more per token but the doc is small (a few thousand chunks).

3. UPGRADED RERANKER: BAAI/bge-reranker-v2-m3
   Ring 3: cross-encoder/ms-marco-MiniLM-L-6-v2 (trained on web search)
   Ring 4: BAAI/bge-reranker-v2-m3 (stronger general-purpose reranker)

   Why: ms-marco MiniLM was trained on web search relevance pairs.
   bge-reranker-v2-m3 scores higher on BEIR and legal text benchmarks.

4. SECTION HEADERS PREPENDED TO CHUNKS
   Ring 3: chunks have page number but no topic label
   Ring 4: detect section headings (ALL-CAPS or Title Case lines) and prepend
            "[Section: Free Zone Persons]" to each chunk during indexing.

   Why: When a chunk says "the threshold is AED 375,000" the embedding doesn't
   know whether this applies to free zones, mainland, or small business relief.
   Section context helps both retrieval (embedding captures topic) and
   generation (LLM knows the rule's scope without inferring it).

5. ANSWER VERIFICATION PASS
   Ring 3: one LLM call → answer (unchecked)
   Ring 4: one LLM call → draft answer
            second LLM call → "Is every claim supported by the context?"
            → if not, strip unsupported claims from the final answer.

   Why: Catches hallucinations (especially Q20's AED penalty amount) before
   they reach the evaluator. Uses gpt-4o-mini (fast, cheap) for this check.

6. STRONGER GENERATION MODEL: gpt-4o
   Ring 3: gpt-4o-mini for all calls
   Ring 4: gpt-4o for final answer generation only
            gpt-4o-mini for query variants, HyDE hypothesis, and verification

   Why: gpt-4o has much better instruction-following and multi-hop reasoning,
   which directly helps Q9-Q12 (complex multi-hop questions). The per-call
   cost difference is small given only 20 questions.

7. FORMULA-NOT-AMOUNT PROMPT RULE (Rule 2 tightened)
   Ring 3: "Never state a specific AED amount unless it appears word-for-word"
   Ring 4: "If the context gives a formula or percentage, state the formula.
            Never compute and state a derived AED amount."

   Why: Q20 (late-filing penalty) has a formula in the doc. LLM was
   computing a specific AED figure and stating it as fact. This rule
   blocks that specific failure mode.

WHAT DID NOT CHANGE FROM RING 3:
─────────────────────────────────────────────────────────────────────────
- Parent-child chunking (parent=2000, child=300)
- Example block preservation (36 blocks, kept whole)
- Multi-query (3 variants per question) — runs alongside HyDE
- Type-aware compression (prose=0.28, example=0.40, pysbd splitting)
- MAX_SENTENCES_PER_CHUNK = 20 cap
- Token budget (3500 tokens, truncate-not-drop, never-empty safety)
- Qdrant in-memory store
- LangSmith observability (env var setup)

Run:
    python ring4_rag.py --pdf "data/CT General Guide - EN - 10 09 2023.pdf" \\
                        --eval uae_tax_rag_eval.xlsx

Outputs:
    ring4_results.md
    ring4_results.json
"""

import argparse
import json
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor

import numpy as np
import tiktoken
from openai import OpenAI
from openpyxl import load_workbook
from langchain_community.document_loaders import PyMuPDFLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
from sentence_transformers import CrossEncoder

# pysbd for robust sentence splitting
try:
    from pysbd import Segmenter as PySBDSegmenter
    _segmenter = PySBDSegmenter(language="en", clean=True)
    def split_sentences(text: str) -> list[str]:
        return [s.strip() for s in _segmenter.segment(text) if s.strip() and len(s.strip()) > 8]
    PYSBD_AVAILABLE = True
except ImportError:
    PYSBD_AVAILABLE = False
    def split_sentences(text: str) -> list[str]:
        raw = re.split(r'(?<=[.!?])\s+(?=[A-Z0-9])', text)
        return [s.strip() for s in raw if s.strip() and len(s.strip()) > 8]

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────────────────────────────────────────
# LangSmith (same as Ring 2/3)
# ─────────────────────────────────────────────────────────────────────────────
LANGSMITH_ENABLED = bool(os.getenv("LANGCHAIN_API_KEY"))
if LANGSMITH_ENABLED:
    os.environ.setdefault("LANGCHAIN_TRACING_V2", "true")
    os.environ.setdefault("LANGCHAIN_PROJECT", "uae-tax-rag")
    print("LangSmith tracing: ON")
else:
    print("LangSmith tracing: OFF — add LANGCHAIN_API_KEY to .env to enable")

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
PARENT_CHUNK_SIZE           = 2000
PARENT_OVERLAP              = 200
CHILD_CHUNK_SIZE            = 300
CHILD_OVERLAP               = 50
RERANKER_MAX_CHARS          = 1500   # slightly larger — bge-reranker handles longer inputs
TOP_K_COSINE                = 30     # per query (question, HyDE, each variant)
TOP_K_RERANK                = 5
NUM_QUERY_VARIANTS          = 3
COMPRESS_THRESHOLD_PROSE    = 0.28
COMPRESS_THRESHOLD_EXAMPLE  = 0.40
MAX_SENTENCES_PER_CHUNK     = 20
MIN_SENTENCES_KEPT          = 3
TOKEN_BUDGET                = 3500

# NEW in Ring 4
EMBED_MODEL      = "text-embedding-3-large"   # upgraded from text-embedding-3-small
CHAT_MODEL       = "gpt-4o"                   # upgraded for final generation
FAST_MODEL       = "gpt-4o-mini"              # kept for variants, HyDE, verification
COLLECTION       = "uae_tax_ring4"
RERANKER_MODEL   = "BAAI/bge-reranker-v2-m3"  # upgraded from ms-marco-MiniLM-L-6-v2

client = OpenAI()
try:
    tokenizer = tiktoken.encoding_for_model("gpt-4o")
except Exception:
    tokenizer = None


def count_tokens(text: str) -> int:
    if tokenizer is not None:
        try:
            return len(tokenizer.encode(text))
        except Exception:
            pass
    return max(1, len(text) // 4)


# ─────────────────────────────────────────────────────────────────────────────
# 1. Section heading detection
# ─────────────────────────────────────────────────────────────────────────────
def detect_section_heading(text: str) -> str | None:
    """
    Looks at the first 2 lines of a chunk and returns the most likely
    section heading if found.

    Heuristics for UAE CT document headings:
      - ALL-CAPS line (e.g. "FREE ZONE PERSONS")
      - Title Case line of 3-12 words that doesn't end with a period
        (e.g. "Small Business Relief")
      - Starts with a numbered section marker (e.g. "4.1 Taxable Income")

    Returns None if no heading detected.
    """
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    for line in lines[:2]:
        # ALL-CAPS heading (ignore very short lines like "A." or "UAE")
        if line.isupper() and 10 < len(line) < 120:
            return line.title()
        # Numbered section like "4.1 Taxable Income" or "Article 12"
        if re.match(r'^(Article\s+\d+|\d+\.\d*)\s+[A-Z]', line):
            return line
        # Title Case, 3-12 words, no trailing period
        words = line.split()
        if (3 <= len(words) <= 12
                and not line.endswith(".")
                and sum(1 for w in words if w[0].isupper()) >= len(words) * 0.6):
            return line
    return None


def prepend_section_header(text: str, chunk_type: str) -> str:
    """Prepend detected section heading so embeddings capture topic context."""
    heading = detect_section_heading(text)
    if heading and chunk_type == "prose":
        return f"[Section: {heading}]\n{text}"
    return text


# ─────────────────────────────────────────────────────────────────────────────
# 2. Build parent + child chunks  (same structure as Ring 3, adds section headers)
# ─────────────────────────────────────────────────────────────────────────────
def build_parent_child_chunks(pdf_path: str) -> tuple[dict, list[dict]]:
    loader = PyMuPDFLoader(pdf_path)
    pages  = loader.load()

    page_texts      = [p.page_content for p in pages]
    page_nums       = [p.metadata.get("page", i) + 1 for i, p in enumerate(pages)]
    cumulative      = 0
    page_boundaries = []
    for text, num in zip(page_texts, page_nums):
        cumulative += len(text) + 1
        page_boundaries.append((cumulative, num))

    full_text = "\n".join(page_texts)

    def estimate_page(char_pos: int) -> int:
        for boundary, page_num in page_boundaries:
            if char_pos <= boundary:
                return page_num
        return page_boundaries[-1][1]

    parts = re.split(r"(Example \d+:)", full_text)

    prose_splitter = RecursiveCharacterTextSplitter(
        chunk_size=PARENT_CHUNK_SIZE,
        chunk_overlap=PARENT_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    child_splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHILD_CHUNK_SIZE,
        chunk_overlap=CHILD_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )

    raw_parents = []
    i = 0
    while i < len(parts):
        part = parts[i]
        if re.match(r"Example \d+:", part):
            content      = parts[i + 1] if i + 1 < len(parts) else ""
            full_example = (part + content).strip()
            if full_example:
                pos = full_text.find(part)
                raw_parents.append({
                    "text": full_example,
                    "page": estimate_page(pos) if pos >= 0 else 1,
                    "type": "example",
                })
            i += 2
        else:
            if part.strip():
                for prose_parent in prose_splitter.split_text(part):
                    if prose_parent.strip():
                        pos = full_text.find(prose_parent[:80])
                        raw_parents.append({
                            "text": prose_parent.strip(),
                            "page": estimate_page(pos) if pos >= 0 else 1,
                            "type": "prose",
                        })
            i += 1

    parent_store: dict[str, dict] = {}
    children:     list[dict]      = []

    for idx, parent in enumerate(raw_parents):
        pid = f"p{idx}"
        # NEW: prepend section header to prose chunks so embeddings capture topic
        indexed_text = prepend_section_header(parent["text"], parent["type"])
        parent_store[pid] = {
            **parent,
            "indexed_text": indexed_text,   # used for child embedding
            "source": os.path.basename(pdf_path),
        }
        for j, child_text in enumerate(child_splitter.split_text(indexed_text)):
            if child_text.strip():
                children.append({
                    "id":        f"c{idx}_{j}",
                    "parent_id": pid,
                    "text":      child_text.strip(),
                    "page":      parent["page"],
                    "source":    os.path.basename(pdf_path),  # Ring 1/2: source in payload
                })

    return parent_store, children


# ─────────────────────────────────────────────────────────────────────────────
# 3. Embed  (upgraded to text-embedding-3-large)
# ─────────────────────────────────────────────────────────────────────────────
def embed_texts(texts: list[str]) -> np.ndarray:
    vectors = []
    BATCH   = 100
    for i in range(0, len(texts), BATCH):
        batch = texts[i : i + BATCH]
        resp  = client.embeddings.create(model=EMBED_MODEL, input=batch)
        vectors.extend(d.embedding for d in resp.data)
    arr   = np.array(vectors, dtype=np.float32)
    norms = np.linalg.norm(arr, axis=1, keepdims=True)
    norms[norms == 0] = 1e-9
    return arr / norms


# ─────────────────────────────────────────────────────────────────────────────
# 4. Qdrant store  (same as Ring 3 — children only)
# ─────────────────────────────────────────────────────────────────────────────
def build_qdrant(children: list[dict], embeddings: np.ndarray) -> QdrantClient:
    qdrant = QdrantClient(":memory:")
    dim    = embeddings.shape[1]
    qdrant.create_collection(
        collection_name=COLLECTION,
        vectors_config=VectorParams(size=dim, distance=Distance.COSINE),
    )
    points = [
        PointStruct(
            id      = i,
            vector  = embeddings[i].tolist(),
            payload = {
                "text":      children[i]["text"],
                "parent_id": children[i]["parent_id"],
                "page":      children[i]["page"],
                "source":    children[i]["source"],   # Ring 1/2: source stored in payload
            },
        )
        for i in range(len(children))
    ]
    qdrant.upsert(collection_name=COLLECTION, points=points)
    return qdrant


# ─────────────────────────────────────────────────────────────────────────────
# 5a. HyDE — generate hypothetical answer for query-side vocabulary bridging
# ─────────────────────────────────────────────────────────────────────────────
HYDE_PROMPT = """You are a UAE Corporate Tax expert. Write a 2-3 sentence answer to the
question below as if you were explaining it from a formal UAE Corporate Tax regulatory
document. Use precise legal and tax terminology.

Question: {question}

Answer (document-style, formal):"""


def generate_hyde_hypothesis(question: str) -> str:
    resp = client.chat.completions.create(
        model=FAST_MODEL,
        temperature=0.0,
        messages=[{
            "role": "user",
            "content": HYDE_PROMPT.format(question=question),
        }],
    )
    return resp.choices[0].message.content.strip()


# ─────────────────────────────────────────────────────────────────────────────
# 5b. Multi-query — same as Ring 3
# ─────────────────────────────────────────────────────────────────────────────
QUERY_EXPANSION_PROMPT = """You are helping retrieve information from a UAE Corporate Tax document.

Given the user's question below, generate {n} alternate phrasings of the same question
using formal legal and tax terminology that might appear in a regulatory document.
Return ONLY the alternate questions, one per line. No numbering, no explanations.

User question: {question}"""


def generate_query_variants(question: str) -> list[str]:
    resp = client.chat.completions.create(
        model=FAST_MODEL,
        temperature=0.3,
        messages=[{
            "role": "user",
            "content": QUERY_EXPANSION_PROMPT.format(
                n=NUM_QUERY_VARIANTS,
                question=question,
            ),
        }],
    )
    raw      = resp.choices[0].message.content.strip()
    variants = [line.strip() for line in raw.split("\n") if line.strip()]
    return variants[:NUM_QUERY_VARIANTS]


# ─────────────────────────────────────────────────────────────────────────────
# 5c. Retrieve children for one query string
# ─────────────────────────────────────────────────────────────────────────────
def retrieve_one(qdrant: QdrantClient, query_vec: np.ndarray) -> list:
    response = qdrant.query_points(
        collection_name=COLLECTION,
        query=query_vec.tolist(),
        limit=TOP_K_COSINE,
        with_payload=True,
    )
    return response.points


# ─────────────────────────────────────────────────────────────────────────────
# 5d. Full retrieval pipeline: HyDE + multi-query → parent lookup → rerank
# ─────────────────────────────────────────────────────────────────────────────
def retrieve_and_rerank(
    qdrant:       QdrantClient,
    reranker:     CrossEncoder,
    question:     str,
    variants:     list[str],
    hyde_text:    str,
    parent_store: dict,
) -> tuple[list[dict], str]:
    """
    Retrieval sources (all run in parallel, deduplicated by parent_id):
      1. Original question embedding
      2. HyDE hypothesis embedding  ← NEW in Ring 4
      3. Each query variant embedding (3 variants)

    After deduplication → rerank using the original question.
    Returns (ranked_parents, hyde_text_for_logging).
    """
    # Embed all queries in one batch to save API round-trips
    all_queries    = [question, hyde_text] + variants
    all_vecs       = embed_texts(all_queries)
    question_vec   = all_vecs[0]
    hyde_vec       = all_vecs[1]
    variant_vecs   = all_vecs[2:]

    # Parallel retrieval across all query vectors
    query_vecs = [question_vec, hyde_vec] + list(variant_vecs)
    with ThreadPoolExecutor(max_workers=len(query_vecs)) as pool:
        futures           = [pool.submit(retrieve_one, qdrant, v) for v in query_vecs]
        results_per_query = [f.result() for f in futures]

    # Parent lookup + deduplication.
    # Ring 2 improvement: when multiple children from the same parent are hit
    # across different queries, track the BEST (max) cosine score rather than
    # the first-seen child's score. This gives a more accurate relevance signal
    # before the reranker runs — the most semantically matching child wins.
    parent_best_score: dict[str, float] = {}
    parent_data:       dict[str, dict]  = {}

    for results in results_per_query:
        for hit in results:
            pid    = hit.payload["parent_id"]
            score  = hit.score
            parent = parent_store.get(pid)
            if not parent:
                continue
            if pid not in parent_data:
                parent_data[pid] = {
                    "parent_id":    pid,
                    "text":         parent["text"],
                    "page":         parent["page"],
                    "type":         parent.get("type", "prose"),
                    "source":       hit.payload.get("source", ""),
                }
                parent_best_score[pid] = score
            else:
                # Update to max cosine score across all children for this parent
                if score > parent_best_score[pid]:
                    parent_best_score[pid] = score

    unique_parents = [
        {**data, "cosine_score": round(parent_best_score[pid], 4)}
        for pid, data in parent_data.items()
    ]

    if not unique_parents:
        return [], hyde_text

    # Rerank against original question using upgraded bge-reranker-v2-m3
    pairs  = [(question, p["text"][:RERANKER_MAX_CHARS]) for p in unique_parents]
    scores = reranker.predict(pairs)

    ranked = sorted(
        zip(unique_parents, scores),
        key=lambda x: x[1],
        reverse=True,
    )[:TOP_K_RERANK]

    return [
        {
            "rank":         i + 1,
            "parent_id":    p["parent_id"],
            "rerank_score": round(float(score), 4),
            "page":         p["page"],
            "type":         p["type"],
            "text":         p["text"],
        }
        for i, (p, score) in enumerate(ranked)
    ], hyde_text


# ─────────────────────────────────────────────────────────────────────────────
# 6. Contextual compression  (same logic as Ring 3 — type-aware, pysbd)
# ─────────────────────────────────────────────────────────────────────────────
def compress_chunks(question: str, chunks: list[dict]) -> list[dict]:
    q_vec      = embed_texts([question])[0]
    compressed = []

    for chunk in chunks:
        chunk_type = chunk.get("type", "prose")
        sentences  = split_sentences(chunk["text"])

        if len(sentences) <= 2:
            compressed.append({**chunk, "compressed": False})
            continue

        threshold = (
            COMPRESS_THRESHOLD_EXAMPLE if chunk_type == "example"
            else COMPRESS_THRESHOLD_PROSE
        )

        sent_vecs = embed_texts(sentences)
        scored    = [
            (sent, float(np.dot(q_vec, vec)))
            for sent, vec in zip(sentences, sent_vecs)
        ]

        kept = [sent for sent, score in scored if score >= threshold]

        if len(kept) < MIN_SENTENCES_KEPT:
            top_sents = {
                s for s, _ in
                sorted(scored, key=lambda x: x[1], reverse=True)[:MIN_SENTENCES_KEPT]
            }
            kept = [sent for sent in sentences if sent in top_sents]

        if len(kept) > MAX_SENTENCES_PER_CHUNK:
            score_map = {sent: score for sent, score in scored}
            top_kept  = sorted(kept, key=lambda s: score_map.get(s, 0), reverse=True)[:MAX_SENTENCES_PER_CHUNK]
            kept      = [s for s in sentences if s in set(top_kept)]

        if not kept:
            compressed.append({**chunk, "compressed": False})
        else:
            compressed.append({
                **chunk,
                "text":       " ".join(kept),
                "compressed": True,
                "kept_sents": len(kept),
                "orig_sents": len(sentences),
            })

    return compressed


# ─────────────────────────────────────────────────────────────────────────────
# 7. Token budget  (same as Ring 3 — truncate-not-drop, never-empty)
# ─────────────────────────────────────────────────────────────────────────────
def apply_token_budget(chunks: list[dict], budget: int) -> list[dict]:
    if not chunks:
        return []

    kept  = []
    total = 0

    for i, chunk in enumerate(chunks):
        tokens = count_tokens(chunk["text"])

        if total + tokens <= budget:
            kept.append(chunk)
            total += tokens
        elif i == 0:
            max_chars = budget * 4
            truncated = {**chunk, "text": chunk["text"][:max_chars] + "…"}
            kept.append(truncated)
            total = budget
            break
        else:
            break

    if not kept and chunks:
        max_chars = budget * 4
        kept = [{**chunks[0], "text": chunks[0]["text"][:max_chars] + "…"}]

    return kept


# ─────────────────────────────────────────────────────────────────────────────
# 8. Generate draft answer  (upgraded to gpt-4o, tighter Rule 2)
# ─────────────────────────────────────────────────────────────────────────────
GROUNDING_PROMPT = """You are a UAE Corporate Tax assistant. Answer using ONLY the context below.

Rules:
1. If the answer depends on conditions, explain those conditions clearly
   using only the context — do NOT say "I don't know" just because the
   answer is not a simple yes/no.
2. Never state a specific AED amount, penalty figure, percentage, or date
   unless it appears word-for-word in the context below.
   If the context provides a formula or rate, state the formula or rate —
   never compute and report a derived monetary figure.
3. If the question is genuinely not covered by the context, say:
   "This is not covered in the provided document."
4. Be concise but complete — include all conditions and thresholds the
   context mentions that are relevant to the question. Do not pad the answer.
5. End every answer with: Source: page [N] of CTGGCT1.

Context:
{context}

Question: {question}
Answer:"""


def generate_draft_answer(context: str, question: str) -> str:
    resp = client.chat.completions.create(
        model=CHAT_MODEL,
        temperature=0,
        messages=[{
            "role": "user",
            "content": GROUNDING_PROMPT.format(context=context, question=question),
        }],
    )
    return resp.choices[0].message.content.strip()


# ─────────────────────────────────────────────────────────────────────────────
# 9. Answer verification pass  (NEW in Ring 4)
# ─────────────────────────────────────────────────────────────────────────────
VERIFICATION_PROMPT = """You are a strict fact-checker for a UAE Corporate Tax assistant.

Below is the retrieved context and a draft answer. Your job:
1. Read each factual claim in the draft answer (amounts, percentages, dates,
   eligibility rules, thresholds, deadlines).
2. Check whether that claim is directly supported by the context.
3. If ALL claims are supported, reply with the draft answer unchanged.
4. If any claim is NOT supported (invented or inferred beyond what the context says),
   rewrite the answer keeping only the supported claims.
   - Keep the Source line unchanged.
   - Do not add new information.
   - Do not change the tone or structure otherwise.

Context:
{context}

Draft answer:
{draft}

Verified answer (return unchanged if fully supported, otherwise corrected):"""


def verify_answer(context: str, question: str, draft: str) -> tuple[str, bool]:
    """
    Returns (verified_answer, was_modified).
    Uses gpt-4o-mini (fast, cheap) — verification doesn't need gpt-4o reasoning.
    """
    resp = client.chat.completions.create(
        model=FAST_MODEL,
        temperature=0,
        messages=[{
            "role": "user",
            "content": VERIFICATION_PROMPT.format(context=context, draft=draft),
        }],
    )
    verified = resp.choices[0].message.content.strip()
    was_modified = (verified.strip() != draft.strip())
    return verified, was_modified


# ─────────────────────────────────────────────────────────────────────────────
# 10. Full answer pipeline
# ─────────────────────────────────────────────────────────────────────────────
def answer_question(
    qdrant:       QdrantClient,
    reranker:     CrossEncoder,
    question:     str,
    parent_store: dict,
) -> dict:
    # Step 1 — parallel: generate query variants AND HyDE hypothesis
    with ThreadPoolExecutor(max_workers=2) as pool:
        variants_future = pool.submit(generate_query_variants, question)
        hyde_future     = pool.submit(generate_hyde_hypothesis, question)
    variants   = variants_future.result()
    hyde_text  = hyde_future.result()

    # Step 2 — HyDE + multi-query → parent lookup → rerank
    retrieved, hyde_logged = retrieve_and_rerank(
        qdrant, reranker, question, variants, hyde_text, parent_store
    )

    # Step 3 — compress parents (type-aware, pysbd)
    compressed = compress_chunks(question, retrieved)

    # Step 4 — token budget
    budgeted = apply_token_budget(compressed, TOKEN_BUDGET)

    # Step 5 — build context string
    context_blocks = [
        f"[page {c['page']}, {c['type']}] {c['text']}"
        for c in budgeted
    ]
    context = "\n\n".join(context_blocks)

    if not context.strip():
        context = "No relevant context was retrieved for this question."

    token_count = count_tokens(context)

    # Step 6 — generate draft answer (gpt-4o)
    draft = generate_draft_answer(context, question)

    # Step 7 — verification pass (gpt-4o-mini checks draft against context)
    final_answer, was_verified = verify_answer(context, question, draft)

    return {
        "answer":        final_answer,
        "draft_answer":  draft,
        "was_verified":  was_verified,
        "retrieved":     retrieved,
        "compressed":    compressed,
        "variants":      variants,
        "hyde_text":     hyde_logged,
        "token_count":   token_count,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def load_questions(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Test Set"]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, question, category = row[0], row[1], row[2]
        if question:
            questions.append({"num": num, "question": question, "category": category})
    wb.close()
    return questions


def write_markdown(results: list[dict], path: str):
    lines = ["# Ring 4 RAG — results to grade", ""]
    lines.append(
        f"Config: embed={EMBED_MODEL}, gen={CHAT_MODEL}, reranker={RERANKER_MODEL}, "
        f"parent={PARENT_CHUNK_SIZE}, child={CHILD_CHUNK_SIZE}, "
        f"cosine_k={TOP_K_COSINE}, rerank_k={TOP_K_RERANK}, "
        f"variants={NUM_QUERY_VARIANTS}, hyde=ON, verify=ON, "
        f"compress_prose={COMPRESS_THRESHOLD_PROSE}/example={COMPRESS_THRESHOLD_EXAMPLE}, "
        f"max_sents={MAX_SENTENCES_PER_CHUNK}, token_budget={TOKEN_BUDGET}, "
        f"pysbd={PYSBD_AVAILABLE}"
    )
    lines.append("")
    # Ring 1 diagnostic tip: classify each failure so you know what to fix next
    lines.append(
        "> **Grading tip (from Ring 1):** For each wrong answer, check the retrieved parents. "
        "If the correct page appears in rank 1-5 but the answer is wrong → **GENERATION failure** (prompt/model issue). "
        "If the correct page is missing entirely → **RETRIEVAL failure** (chunking/embedding/reranking issue). "
        "Tag each Q before running the evaluator."
    )
    lines.append("")
    for r in results:
        lines.append(f"## Q{r['num']} — {r['category']}")
        lines.append(f"**Question:** {r['question']}")
        lines.append("")
        lines.append(f"**HyDE hypothesis:** {r.get('hyde_text', '?')}")
        lines.append("")
        lines.append("**Query variants:**")
        for v in r.get("variants", []):
            lines.append(f"- {v}")
        lines.append("")
        lines.append(f"**Tokens in context:** {r.get('token_count', '?')}")
        lines.append(f"**Verification modified answer:** {r.get('was_verified', False)}")
        lines.append("")
        if r.get("was_verified"):
            lines.append(f"**Draft answer:** {r.get('draft_answer', '')}")
            lines.append("")
        lines.append(f"**Final answer:** {r['answer']}")
        lines.append("")
        lines.append("**Retrieved parents (after reranking + compression):**")
        for c in r.get("compressed", []):
            preview = c["text"].replace("\n", " ")
            if len(preview) > 300:
                preview = preview[:300] + "…"
            comp_note = (
                f"compressed {c['orig_sents']}→{c['kept_sents']} sentences"
                if c.get("compressed") else "kept as-is"
            )
            lines.append(
                f"- [rank {c['rank']}] page {c['page']} [{c['type']}] "
                f"(rerank {c['rerank_score']}, {comp_note}) {preview}"
            )
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf",  required=True)
    ap.add_argument("--eval", required=True)
    args = ap.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        sys.exit("Set OPENAI_API_KEY first.")

    print("Building parent + child chunks (with section headers)...")
    parent_store, children = build_parent_child_chunks(args.pdf)
    prose_count   = sum(1 for p in parent_store.values() if p["type"] == "prose")
    example_count = sum(1 for p in parent_store.values() if p["type"] == "example")
    headed_count  = sum(
        1 for p in parent_store.values()
        if p.get("indexed_text", "") != p.get("text", "")
    )
    print(f"  {len(parent_store)} parents ({prose_count} prose, {example_count} examples)")
    print(f"  {headed_count} prose parents got section headers prepended")
    print(f"  {len(children)} children (embedded with {EMBED_MODEL})")

    print(f"\nEmbedding children with {EMBED_MODEL}...")
    child_texts = [c["text"] for c in children]
    embeddings  = embed_texts(child_texts)

    print("Building Qdrant store...")
    qdrant = build_qdrant(children, embeddings)

    print(f"Loading reranker ({RERANKER_MODEL})...")
    reranker = CrossEncoder(RERANKER_MODEL)

    print("Loading questions...")
    questions = load_questions(args.eval)
    print(f"  {len(questions)} questions\n")

    results = []
    for q in questions:
        print(f"  Q{q['num']}: {q['question'][:55]}...")
        out = answer_question(qdrant, reranker, q["question"], parent_store)
        tok      = out.get("token_count", "?")
        verified = "✓ verified clean" if not out.get("was_verified") else "⚠ verification modified"
        print(f"         → {tok} tokens | {verified}")
        results.append({**q, **out})

    write_markdown(results, "ring4_results.md")
    with open("ring4_results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    modified_count = sum(1 for r in results if r.get("was_verified"))
    print(f"\nDone. Verification caught hallucinations in {modified_count}/{len(results)} answers.")
    print("Key things to check in ring4_results.md:")
    print("  1. HyDE hypothesis — does it look like document-style language?")
    print("  2. Q20 — did verification strip the invented penalty amount?")
    print("  3. Q9-Q12 — are multi-hop answers more complete (gpt-4o effect)?")
    print("  4. Section headers in rank-1 chunks — do they name the right topic?")
    print("\nThen add 'Ring 4': 'ring4_results.json' to evaluate_tax_v2.py RESULT_FILES and run it.")


if __name__ == "__main__":
    main()
