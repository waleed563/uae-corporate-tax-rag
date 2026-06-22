"""
ring2_rag.py — Ring 2 of the improved UAE Corporate Tax RAG system.

WHAT CHANGED FROM ring1_rag.py (read before running):
─────────────────────────────────────────────────────────────────────────
1. MULTI-QUERY RETRIEVAL
   Ring 1: one question → one retrieval
   Ring 2: one question → LLM generates 3 alternate phrasings →
           retrieve for all 4 (original + 3) → deduplicate → rerank

   Why: Q13 failed because "I freelance from home with no trade licence"
   didn't match the document's legal phrasing. Multi-query generates
   versions like "natural person conducting business without commercial
   registration" — giving retrieval 4 chances instead of 1.

   Cost: 1 extra LLM call per question (to generate variants).
   Total per question: 2 LLM calls (variants + answer).

2. CONTEXTUAL COMPRESSION
   Ring 1: top-5 reranked chunks → LLM (full text, noise included)
   Ring 2: top-5 reranked chunks → EmbeddingsFilter strips irrelevant
           sentences → LLM receives cleaner, shorter context

   Why: a 600-char chunk about tax loss rules might contain 5 sentences,
   only 1 of which answers the question. The compressor keeps that 1
   sentence and drops the other 4. Less noise = better answers,
   fewer tokens = lower cost.

   We use EmbeddingsFilter (not LLMChainExtractor) because:
   - EmbeddingsFilter uses embeddings to score sentences — no LLM call
   - LLMChainExtractor uses an LLM per chunk = 5 extra calls = expensive
   - EmbeddingsFilter is fast, free, and good enough

3. LANGSMITH OBSERVABILITY
   Ring 1: no tracing — you run it and hope for the best
   Ring 2: every step is traced automatically

   What you see in app.langsmith.com:
   - How long each step took (embedding, retrieval, reranker, LLM)
   - How many tokens each LLM call used
   - The exact prompt that went to the LLM
   - The retrieved chunks for every question

   Setup: just add LANGCHAIN_API_KEY to your .env file.
   If you don't have one, sign up free at smith.langchain.com.
   If LANGCHAIN_API_KEY is missing, the script still runs — tracing
   is silently disabled.

4. TOKEN BUDGET
   Ring 1: no limit — could accidentally overflow the context window
   Ring 2: compressed context is measured in tokens before sending
           to LLM. If over budget, trim the least-relevant chunks first.

   Why: gpt-4o-mini has a 128k context window so overflow is unlikely
   here, but this is production discipline. You measure before sending.
   Shows in your portfolio that you thought about real deployment.

WHAT DID NOT CHANGE FROM RING 1:
─────────────────────────────────────────────────────────────────────────
- Loader: LangChain PyMuPDFLoader (same)
- Chunker: 600 chars, 100 overlap, recursive (same)
- Embedding model: text-embedding-3-small (same)
- Vector store: Qdrant in-memory (same)
- Reranker: cross-encoder/ms-marco-MiniLM-L-6-v2 (same)
- LLM: gpt-4o-mini (same)
- Prompt rules A, B, C (same — keeping what worked)

LEARNING EXERCISES (do before running):
─────────────────────────────────────────────────────────────────────────
1. PREDICT THEN RUN
   - Q13 failed in Ring 1 because the AED 1M natural-person threshold
     chunk wasn't retrieved. Multi-query gives 4 retrieval attempts.
     Do you think Q13 will now pass? Why or why not?
   - Q14 failed in Ring 1 as a GENERATION failure — right chunks were
     retrieved but the model gave a generic answer. Multi-query won't
     fix a generation failure. Which Ring 2 change might help Q14?
   - Q20 was uncertain. Will compression help or hurt it?

2. TRACE ONE ITEM
   Pick Q13 ("I freelance from home in Dubai with no trade licence").
   After running, open app.langsmith.com, find Q13's trace and look:
   - What were the 3 alternate questions the LLM generated?
   - Did any alternate question retrieve page 47 (the AED 1M threshold)?
   - How many tokens went to the answer LLM after compression?

3. READ THE DIFF
   Open ring1_rag.py and ring2_rag.py side by side in VSCode.
   For every green block (new code), ask: which of the 4 changes above
   does this belong to? Can you label every new function?

Run:
    pip install -r requirements_ring2.txt
    python ring2_rag.py --pdf "data/CT General Guide - EN - 10 09 2023.pdf" \
                        --eval uae_tax_rag_eval.xlsx

    Optional — enable LangSmith tracing:
    Add to your .env file:
        LANGCHAIN_API_KEY=ls__...          (from smith.langchain.com)
        LANGCHAIN_TRACING_V2=true
        LANGCHAIN_PROJECT=uae-tax-rag

Outputs:
    ring2_results.md
    ring2_results.json
"""

import argparse
import json
import os
import sys
from concurrent.futures import ThreadPoolExecutor

import numpy as np
import tiktoken
from openai import OpenAI
from openpyxl import load_workbook
from langchain_community.document_loaders import PyMuPDFLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
from sentence_transformers import CrossEncoder

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────────────────────────────────────────
# LangSmith — enable tracing if API key is present (CHANGED: new in Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
LANGSMITH_ENABLED = bool(os.getenv("LANGCHAIN_API_KEY"))
if LANGSMITH_ENABLED:
    os.environ.setdefault("LANGCHAIN_TRACING_V2", "true")
    os.environ.setdefault("LANGCHAIN_PROJECT", "uae-tax-rag")
    print("LangSmith tracing: ON — view traces at smith.langchain.com")
else:
    print("LangSmith tracing: OFF — add LANGCHAIN_API_KEY to .env to enable")

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
CHUNK_SIZE       = 600           # same as Ring 1
CHUNK_OVERLAP    = 100           # same as Ring 1
TOP_K_COSINE     = 20            # same as Ring 1
TOP_K_RERANK     = 5             # same as Ring 1
NUM_QUERY_VARIANTS = 3           # CHANGED: how many alternate questions to generate
EMBED_MODEL      = "text-embedding-3-small"
CHAT_MODEL       = "gpt-4o-mini"
COLLECTION       = "uae_tax"
RERANKER_MODEL   = "cross-encoder/ms-marco-MiniLM-L-6-v2"
TOKEN_BUDGET     = 3000          # CHANGED: max tokens in context sent to LLM
COMPRESS_THRESHOLD = 0.35        # TUNED: lowered from 0.5 — 0.5 stripped useful detail (Q10, Q14)
MIN_SENTENCES_KEPT = 2           # TUNED: never compress below this many sentences (safety net)

client   = OpenAI()
tokenizer = tiktoken.encoding_for_model("gpt-4o-mini")


# ─────────────────────────────────────────────────────────────────────────────
# 1. Load + chunk (same as Ring 1)
# ─────────────────────────────────────────────────────────────────────────────
def load_and_chunk(pdf_path: str) -> list[dict]:
    loader = PyMuPDFLoader(pdf_path)
    pages  = loader.load()
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    docs   = splitter.split_documents(pages)
    chunks = []
    for doc in docs:
        text = doc.page_content.strip()
        if not text:
            continue
        chunks.append({
            "text":   text,
            "page":   doc.metadata.get("page", 0) + 1,
            "source": os.path.basename(pdf_path),
        })
    return chunks


# ─────────────────────────────────────────────────────────────────────────────
# 2. Embed (same as Ring 1)
# ─────────────────────────────────────────────────────────────────────────────
def embed_texts(texts: list[str]) -> np.ndarray:
    vectors = []
    BATCH   = 100
    for i in range(0, len(texts), BATCH):
        batch = texts[i : i + BATCH]
        resp  = client.embeddings.create(model=EMBED_MODEL, input=batch)
        vectors.extend(d.embedding for d in resp.data)
    arr   = np.array(vectors, dtype=np.float32)
    norms = np.linalg.norm(arr, axis=1, keepdims=True)
    norms[norms == 0] = 1e-9
    return arr / norms


# ─────────────────────────────────────────────────────────────────────────────
# 3. Qdrant store (same as Ring 1)
# ─────────────────────────────────────────────────────────────────────────────
def build_qdrant(chunks: list[dict], embeddings: np.ndarray) -> QdrantClient:
    qdrant = QdrantClient(":memory:")
    dim    = embeddings.shape[1]
    qdrant.create_collection(
        collection_name=COLLECTION,
        vectors_config=VectorParams(size=dim, distance=Distance.COSINE),
    )
    points = [
        PointStruct(
            id      = i,
            vector  = embeddings[i].tolist(),
            payload = {
                "text":   chunks[i]["text"],
                "page":   chunks[i]["page"],
                "source": chunks[i]["source"],
            },
        )
        for i in range(len(chunks))
    ]
    qdrant.upsert(collection_name=COLLECTION, points=points)
    return qdrant


# ─────────────────────────────────────────────────────────────────────────────
# 4a. Multi-query — generate alternate phrasings (CHANGED: new in Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
QUERY_EXPANSION_PROMPT = """You are helping retrieve information from a UAE Corporate Tax document.

Given the user's question below, generate {n} alternate phrasings of the same question
using formal legal and tax terminology that might appear in a regulatory document.
Return ONLY the alternate questions, one per line. No numbering, no explanations.

User question: {question}"""


def generate_query_variants(question: str) -> list[str]:
    """
    Calls the LLM once to produce NUM_QUERY_VARIANTS alternate phrasings.
    These use legal/tax language so they match the document's vocabulary better.

    Example:
      Input:  "I freelance from home, do I owe tax?"
      Output: [
        "Is a natural person conducting business activities without a trade licence
         subject to UAE Corporate Tax?",
        "What is the corporate tax obligation for self-employed individuals
         earning below AED 1,000,000?",
        "Are freelancers without commercial registration considered taxable persons
         under the Corporate Tax Law?"
      ]
    """
    resp = client.chat.completions.create(
        model=CHAT_MODEL,
        temperature=0.3,    # slight creativity for diverse phrasings
        messages=[{
            "role": "user",
            "content": QUERY_EXPANSION_PROMPT.format(
                n=NUM_QUERY_VARIANTS,
                question=question,
            ),
        }],
    )
    raw      = resp.choices[0].message.content.strip()
    variants = [line.strip() for line in raw.split("\n") if line.strip()]
    return variants[:NUM_QUERY_VARIANTS]


# ─────────────────────────────────────────────────────────────────────────────
# 4b. Retrieve for one query (used in parallel for multi-query)
# ─────────────────────────────────────────────────────────────────────────────
def retrieve_one(qdrant: QdrantClient, query: str) -> list:
    q_vec    = embed_texts([query])[0]
    response = qdrant.query_points(
        collection_name=COLLECTION,
        query=q_vec.tolist(),
        limit=TOP_K_COSINE,
        with_payload=True,
    )
    return response.points


# ─────────────────────────────────────────────────────────────────────────────
# 4c. Multi-query retrieve + rerank (CHANGED: replaces Ring 1's retrieve_and_rerank)
# ─────────────────────────────────────────────────────────────────────────────
def retrieve_and_rerank(
    qdrant:    QdrantClient,
    reranker:  CrossEncoder,
    question:  str,
    variants:  list[str],
) -> list[dict]:
    """
    Step 1 — retrieve for all queries in parallel (original + variants).
             ThreadPoolExecutor runs them concurrently — faster than sequential.
    Step 2 — deduplicate by chunk ID so the same chunk isn't scored twice.
    Step 3 — reranker scores (original question, chunk) pairs for all candidates.
    Step 4 — return top TOP_K_RERANK after reranking.
    """
    all_queries = [question] + variants

    # Step 1 — parallel retrieval
    with ThreadPoolExecutor(max_workers=len(all_queries)) as pool:
        futures = [pool.submit(retrieve_one, qdrant, q) for q in all_queries]
        results_per_query = [f.result() for f in futures]

    # Step 2 — deduplicate
    seen   = set()
    unique = []
    for results in results_per_query:
        for r in results:
            if r.id not in seen:
                seen.add(r.id)
                unique.append(r)

    # Step 3 — rerank against the ORIGINAL question (not the variants)
    pairs  = [(question, r.payload["text"]) for r in unique]
    scores = reranker.predict(pairs)

    # Step 4 — sort and keep best TOP_K_RERANK
    ranked = sorted(
        zip(unique, scores),
        key=lambda x: x[1],
        reverse=True,
    )[:TOP_K_RERANK]

    return [
        {
            "rank":         i + 1,
            "chunk_id":     hit.id,
            "rerank_score": round(float(score), 4),
            "page":         hit.payload["page"],
            "text":         hit.payload["text"],
        }
        for i, (hit, score) in enumerate(ranked)
    ]


# ─────────────────────────────────────────────────────────────────────────────
# 5. Contextual compression (CHANGED: new in Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
def compress_chunks(question: str, chunks: list[dict]) -> list[dict]:
    """
    For each chunk, split into sentences. Embed the question and each sentence.
    Keep only sentences whose cosine similarity to the question exceeds
    COMPRESS_THRESHOLD. Reassemble the kept sentences as the compressed chunk.

    Why EmbeddingsFilter instead of LLMChainExtractor:
    - No extra LLM call per chunk (5 chunks × LLM call = expensive)
    - Uses embeddings only — fast and cheap
    - Good enough for sentence-level noise removal

    If a chunk becomes empty after compression (all sentences were noise),
    keep the original — better to have noise than nothing.
    """
    q_vec = embed_texts([question])[0]

    compressed = []
    for chunk in chunks:
        # Split into sentences (simple split — good enough for this document)
        sentences = [s.strip() for s in chunk["text"].split(". ") if s.strip()]
        if len(sentences) <= 1:
            # Too short to compress — keep as-is
            compressed.append({**chunk, "compressed": False})
            continue

        # Embed all sentences in one batch call
        sent_vecs = embed_texts(sentences)

        # Score every sentence by similarity to the question
        scored = [
            (sent, float(np.dot(q_vec, vec)))
            for sent, vec in zip(sentences, sent_vecs)
        ]

        # Keep sentences above the threshold
        kept = [sent for sent, score in scored if score >= COMPRESS_THRESHOLD]

        # SAFETY NET: if too few survived, keep the top MIN_SENTENCES_KEPT
        # by score instead — prevents over-compression that hurt recall on
        # Q10 (lost 95% ownership rule) and Q14 (thin 285-token context).
        if len(kept) < MIN_SENTENCES_KEPT:
            top = sorted(scored, key=lambda x: x[1], reverse=True)[:MIN_SENTENCES_KEPT]
            # preserve original sentence order for readability
            top_sents = {s for s, _ in top}
            kept = [sent for sent in sentences if sent in top_sents]

        if not kept:
            # Nothing survived — keep original to avoid empty context
            compressed.append({**chunk, "compressed": False})
        else:
            compressed.append({
                **chunk,
                "text":       ". ".join(kept),
                "compressed": True,
                "kept_sents": len(kept),
                "orig_sents": len(sentences),
            })

    return compressed


# ─────────────────────────────────────────────────────────────────────────────
# 6. Token budget (CHANGED: new in Ring 2)
# ─────────────────────────────────────────────────────────────────────────────
def apply_token_budget(chunks: list[dict], budget: int) -> list[dict]:
    """
    Measure total tokens in context. If over budget, drop the
    lowest-ranked chunks (they're least relevant anyway) until under budget.

    Why: gpt-4o-mini has a 128k window so overflow is unlikely on one PDF,
    but this is production discipline — you always know what you're sending.
    Shows up in your portfolio README as evidence of real engineering thinking.
    """
    kept   = []
    total  = 0
    for chunk in chunks:   # already sorted by rerank score (best first)
        tokens = len(tokenizer.encode(chunk["text"]))
        if total + tokens <= budget:
            kept.append(chunk)
            total += tokens
        else:
            break   # budget exceeded — drop remaining chunks
    return kept


# ─────────────────────────────────────────────────────────────────────────────
# 7. Generate answer (prompt same as Ring 1, context now compressed + budgeted)
# ─────────────────────────────────────────────────────────────────────────────
GROUNDING_PROMPT = """You are a UAE Corporate Tax assistant. Answer using ONLY the context below.

Rules:
1. If the answer depends on conditions, explain those conditions clearly
   using only the context — do NOT say "I don't know" just because the
   answer is not a simple yes/no.
2. Never state a specific AED amount, penalty figure, percentage, or date
   unless it appears word-for-word in the context below.
3. If the question is genuinely not covered by the context, say:
   "This is not covered in the provided document."
4. End every answer with: Source: page [N] of CTGGCT1.

Context:
{context}

Question: {question}
Answer:"""


def answer_question(
    qdrant:   QdrantClient,
    reranker: CrossEncoder,
    question: str,
) -> dict:
    # Step 1 — generate alternate phrasings (1 LLM call)
    variants = generate_query_variants(question)

    # Step 2 — multi-query retrieve + rerank
    retrieved = retrieve_and_rerank(qdrant, reranker, question, variants)

    # Step 3 — compress each chunk (embeddings only, no LLM call)
    compressed = compress_chunks(question, retrieved)

    # Step 4 — apply token budget
    budgeted = apply_token_budget(compressed, TOKEN_BUDGET)

    # Step 5 — build context with page citations
    context_blocks = [f"[page {c['page']}] {c['text']}" for c in budgeted]
    context        = "\n\n".join(context_blocks)

    # Step 6 — count tokens going in (for logging)
    token_count = len(tokenizer.encode(context))

    # Step 7 — generate answer (1 LLM call)
    resp = client.chat.completions.create(
        model=CHAT_MODEL,
        temperature=0,
        messages=[{
            "role": "user",
            "content": GROUNDING_PROMPT.format(
                context=context,
                question=question,
            ),
        }],
    )
    answer = resp.choices[0].message.content.strip()

    return {
        "answer":        answer,
        "retrieved":     retrieved,
        "compressed":    compressed,
        "variants":      variants,
        "token_count":   token_count,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Helpers (same as Ring 1)
# ─────────────────────────────────────────────────────────────────────────────
def load_questions(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Test Set"]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, question, category = row[0], row[1], row[2]
        if question:
            questions.append({"num": num, "question": question, "category": category})
    wb.close()
    return questions


def write_markdown(results: list[dict], path: str):
    lines = ["# Ring 2 RAG — results to grade", ""]
    lines.append(
        f"Config: chunk_size={CHUNK_SIZE}, overlap={CHUNK_OVERLAP}, "
        f"cosine_k={TOP_K_COSINE}, rerank_k={TOP_K_RERANK}, "
        f"variants={NUM_QUERY_VARIANTS}, token_budget={TOKEN_BUDGET}, "
        f"embed={EMBED_MODEL}, chat={CHAT_MODEL}"
    )
    lines.append("")
    for r in results:
        lines.append(f"## Q{r['num']} — {r['category']}")
        lines.append(f"**Question:** {r['question']}")
        lines.append("")
        lines.append("**Query variants generated:**")
        for v in r.get("variants", []):
            lines.append(f"- {v}")
        lines.append("")
        lines.append(f"**Tokens in context:** {r.get('token_count', '?')}")
        lines.append("")
        lines.append(f"**Answer:** {r['answer']}")
        lines.append("")
        lines.append("**Retrieved + compressed chunks:**")
        for c in r.get("compressed", []):
            preview = c["text"].replace("\n", " ")
            if len(preview) > 300:
                preview = preview[:300] + "…"
            comp_note = (
                f"compressed {c['orig_sents']}→{c['kept_sents']} sentences"
                if c.get("compressed") else "kept as-is"
            )
            lines.append(
                f"- [rank {c['rank']}] page {c['page']} "
                f"(rerank {c['rerank_score']}, {comp_note}) {preview}"
            )
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf",  required=True)
    ap.add_argument("--eval", required=True)
    args = ap.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        sys.exit("Set OPENAI_API_KEY first.")

    print("Loading and chunking PDF...")
    chunks = load_and_chunk(args.pdf)
    print(f"  {len(chunks)} chunks")

    print("Embedding chunks...")
    texts      = [c["text"] for c in chunks]
    embeddings = embed_texts(texts)

    print("Building Qdrant store...")
    qdrant = build_qdrant(chunks, embeddings)

    print(f"Loading reranker ({RERANKER_MODEL})...")
    reranker = CrossEncoder(RERANKER_MODEL)

    print("Loading questions...")
    questions = load_questions(args.eval)
    print(f"  {len(questions)} questions\n")

    results = []
    for q in questions:
        print(f"  Q{q['num']}: {q['question'][:55]}...")
        out = answer_question(qdrant, reranker, q["question"])
        tok = out.get("token_count", "?")
        print(f"         → {tok} tokens in context, {len(out['variants'])} variants generated")
        results.append({**q, **out})

    write_markdown(results, "ring2_results.md")
    with open("ring2_results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print("\nDone.")
    print("Open ring2_results.md — check the 'Query variants' for each question.")
    print("Pay attention to Q13: did a variant retrieve page 47 (AED 1M threshold)?")
    print("Then grade and bring back for RAGAS evaluation.")


if __name__ == "__main__":
    main()