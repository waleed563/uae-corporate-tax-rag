"""
fill_answer_key.py — Automatically fills column E of uae_tax_rag_eval.xlsx
                     by asking GPT to extract the correct answer from the
                     full UAE Corporate Tax General Guide PDF.

HOW IT WORKS:
─────────────────────────────────────────────────────────────────────────
The PDF is 270k characters (~67k tokens). GPT-4o-mini has a 128k token
window. So we send the ENTIRE document + each question in one call.
No retrieval, no chunking — the LLM reads the full document and extracts
the precise answer directly.

This is the most accurate way to build a ground-truth answer key.

For trap questions (VAT, Saudi Arabia, salary income tax) — the LLM
recognises they are not in the document and writes "NOT COVERED", which
the evaluator (evaluate_tax_v2.py) handles correctly.

RUN (main tax-rag env):
─────────────────────────────────────────────────────────────────────────
    python fill_answer_key.py \\
        --pdf "data/CT General Guide - EN - 10 09 2023.pdf" \\
        --eval uae_tax_rag_eval.xlsx

The script writes answers into column E and saves the file.
It skips any row that already has an answer key filled (safe to re-run).
"""

import argparse
import os
import sys
import time

import fitz                          # PyMuPDF
from openai import OpenAI
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

client = OpenAI()

ANSWER_KEY_PROMPT = """You are a UAE Corporate Tax expert reading the official
Federal Tax Authority Corporate Tax General Guide (CTGGCT1).

The FULL document text is provided below.

Your job: answer the question below using ONLY what is written in this document.

Rules:
1. Be precise and concise — 1-3 sentences maximum.
2. Include specific figures, thresholds, or conditions if they appear.
3. If the question asks about something NOT covered in this document
   (e.g. VAT, other countries' taxes, personal salary income tax),
   reply with exactly: NOT COVERED
4. Do not add any explanation, preamble, or "Source:" line.
   Just the answer.

Question: {question}

Document:
{document}

Answer:"""


# ─────────────────────────────────────────────────────────────────────────────
# Extract full PDF text
# ─────────────────────────────────────────────────────────────────────────────
def load_pdf(path: str) -> str:
    doc = fitz.open(path)
    pages = [page.get_text() for page in doc]
    doc.close()
    return "\n".join(pages)


# ─────────────────────────────────────────────────────────────────────────────
# Ask GPT to extract the answer for one question from the full document
# ─────────────────────────────────────────────────────────────────────────────
def get_answer(question: str, document: str) -> str:
    prompt = ANSWER_KEY_PROMPT.format(
        question=question,
        document=document,
    )
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        temperature=0,
        messages=[{"role": "user", "content": prompt}],
    )
    return resp.choices[0].message.content.strip()


# ─────────────────────────────────────────────────────────────────────────────
# Load questions + existing answers from xlsx
# ─────────────────────────────────────────────────────────────────────────────
def load_questions(xlsx_path: str) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Test Set"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num      = row[0]
        question = row[1]
        category = row[2]
        existing = row[4] if len(row) > 4 else None   # column E
        if num and question:
            rows.append({
                "num":      num,
                "question": question,
                "category": category,
                "existing": (existing or "").strip(),
            })
    wb.close()
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Write answers back into column E of the xlsx
# ─────────────────────────────────────────────────────────────────────────────
def write_answers(xlsx_path: str, answers: dict[int, str]):
    """answers = {question_num: answer_text}"""
    wb = load_workbook(xlsx_path)
    ws = wb["Test Set"]
    for row in ws.iter_rows(min_row=2):
        num_cell    = row[0]    # column A
        answer_cell = row[4]    # column E
        if num_cell.value in answers:
            answer_cell.value = answers[num_cell.value]
    wb.save(xlsx_path)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf",  required=True, help="Path to CTGGCT1 PDF")
    ap.add_argument("--eval", required=True, help="Path to uae_tax_rag_eval.xlsx")
    ap.add_argument("--overwrite", action="store_true",
                    help="Overwrite existing answers (default: skip filled rows)")
    args = ap.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        sys.exit("Set OPENAI_API_KEY first.")

    print("Loading PDF (full text)...")
    document = load_pdf(args.pdf)
    print(f"  {len(document):,} characters loaded")
    print(f"  Estimated tokens: ~{len(document)//4:,} (fits in gpt-4o-mini 128k window)")

    print("\nLoading questions from spreadsheet...")
    questions = load_questions(args.eval)
    print(f"  {len(questions)} questions found")

    to_fill = [q for q in questions
               if not q["existing"] or args.overwrite]
    skipped = len(questions) - len(to_fill)
    if skipped:
        print(f"  {skipped} already filled — will skip (use --overwrite to redo)")
    print(f"  {len(to_fill)} questions to fill\n")

    answers = {}
    for i, q in enumerate(to_fill, 1):
        print(f"[{i}/{len(to_fill)}] Q{q['num']} — {q['question'][:60]}...")
        answer = get_answer(q["question"], document)
        answers[q["num"]] = answer
        tag = "NOT COVERED" if answer == "NOT COVERED" else answer[:80] + "…"
        print(f"         → {tag}")
        time.sleep(0.5)     # gentle on rate limits

    print("\nWriting answers to spreadsheet...")
    write_answers(args.eval, answers)
    print(f"  Saved — {len(answers)} answers written to column E")
    print("\nDone. Open the spreadsheet to review, then run:")
    print("  python evaluate_tax_v2.py")


if __name__ == "__main__":
    main()