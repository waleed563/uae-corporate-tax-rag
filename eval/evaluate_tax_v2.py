"""
evaluate_tax_v2.py — Upgraded LLM-as-judge for the UAE Tax RAG project.

WHAT CHANGED FROM evaluate_tax.py (and WHY your scores didn't match the answers):
─────────────────────────────────────────────────────────────────────────
1. USES THE ANSWER KEY (ground truth) from the spreadsheet.
   The old version judged with NO correct answer to compare against, so
   precision and recall were guesses. Now the judge sees the real answer
   from the document and grades against it. This is the main fix.

2. TRAP-AWARE judging.
   Old version scored a correct refusal ("not in this document") as 0.0 —
   punishing the system for doing the RIGHT thing on VAT/Saudi/salary
   questions. Now: if the answer key says "not covered", a refusal scores
   1.0 and a made-up answer scores 0.0. Correct behaviour is rewarded.

3. STRICTER faithfulness.
   Old version gave everything 1.00 (too lenient). Now the judge is told
   to drop the score if ANY specific number, date, or claim is not
   supported by the context — even one. Catches the cases you can see
   but the old judge missed.

REQUIREMENT:
─────────────────────────────────────────────────────────────────────────
Column E ("Answer key") in uae_tax_rag_eval.xlsx MUST be filled.
If a row's answer key is empty, that question is SKIPPED (not guessed)
and reported separately, so your averages stay honest.

For trap questions (Q17-Q20), write the answer key as exactly:
    NOT COVERED
(or start the line with "not covered" / "out of scope" — any of these work)
This tells the judge the correct behaviour is to refuse.

RUN (main tax-rag env):
    python evaluate_tax_v2.py
"""

import json
import time
from openai import OpenAI
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

client = OpenAI()

RESULT_FILES = {
    "Baseline": "baseline_results.json",
    "Ring 1":   "ring1_results.json",
    "Ring 2":   "ring2_results.json",
    "Ring 3":   "ring3_results.json",
    "Ring 4":   "ring4_results.json",
}
EVAL_XLSX = "uae_tax_rag_eval.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# Load answer key from column E
# ─────────────────────────────────────────────────────────────────────────────
def load_answer_key(xlsx_path: str) -> dict[int, str]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Test Set"]
    keys = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        num = row[0]
        answer_key = row[4] if len(row) > 4 else None
        if num:
            keys[num] = (answer_key or "").strip()
    wb.close()
    return keys


def is_trap(answer_key: str) -> bool:
    k = answer_key.lower()
    return k.startswith("not covered") or k.startswith("out of scope") or k == "not covered"


# ─────────────────────────────────────────────────────────────────────────────
# Judge for a NORMAL (answerable) question — graded against ground truth
# ─────────────────────────────────────────────────────────────────────────────
def score_normal(question: str, answer: str, context: str, ground_truth: str) -> dict:
    judge_prompt = f"""
You are strictly evaluating a RAG system that answers UAE Corporate Tax questions.
You are given the CORRECT answer (ground truth) from the official document.
Return ONLY a JSON object, nothing else.

Question: {question}

CORRECT answer (ground truth from the document):
{ground_truth}

Context the system retrieved:
{context[:3000]}

Answer the system gave:
{answer[:1500]}

Score each metric from 0.0 to 1.0. Be strict and use the full range:

{{
  "faithfulness": <Is EVERY claim, number, and date in the system's answer supported by the retrieved context? If even ONE specific figure or claim is unsupported, score 0.5 or lower. 1.0 only if fully grounded.>,
  "answer_relevancy": <Does the system's answer match the CORRECT answer above and directly address the question? 1.0=matches the correct answer well, 0.0=wrong or off-topic.>,
  "context_precision": <Were the retrieved chunks actually relevant to producing the correct answer? 1.0=all useful, 0.0=mostly noise.>,
  "context_recall": <Did the retrieved context contain the information present in the CORRECT answer? 1.0=all key facts present, 0.0=key facts missing.>
}}
"""
    return _call_judge(judge_prompt)


# ─────────────────────────────────────────────────────────────────────────────
# Judge for a TRAP question — correct behaviour is to refuse
# ─────────────────────────────────────────────────────────────────────────────
def score_trap(question: str, answer: str) -> dict:
    judge_prompt = f"""
You are evaluating a RAG system on an OUT-OF-SCOPE question.
This question is NOT answerable from the UAE Corporate Tax General Guide.
The CORRECT behaviour is for the system to REFUSE — to say it doesn't know,
or that this is not covered in the document.

Return ONLY a JSON object, nothing else.

Question: {question}

Answer the system gave:
{answer[:1500]}

Score:
{{
  "refused_correctly": <Did the system correctly refuse / say it is not covered, instead of inventing an answer? 1.0=correctly refused, 0.0=made up an answer>
}}
"""
    raw = _call_judge_raw(judge_prompt)
    try:
        clean = raw.replace("```json", "").replace("```", "").strip()
        val = json.loads(clean).get("refused_correctly", 0.0)
    except Exception:
        val = 0.0
    # Map the single trap score onto all four metrics so it averages cleanly
    return {
        "faithfulness":      val,
        "answer_relevancy":  val,
        "context_precision": val,
        "context_recall":    val,
        "_trap":             True,
        "_refused":          val,
    }


def _call_judge(prompt: str) -> dict:
    raw = _call_judge_raw(prompt)
    try:
        clean = raw.replace("```json", "").replace("```", "").strip()
        return json.loads(clean)
    except Exception:
        return {"faithfulness": 0.0, "answer_relevancy": 0.0,
                "context_precision": 0.0, "context_recall": 0.0}


def _call_judge_raw(prompt: str) -> str:
    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0,
    )
    return resp.choices[0].message.content.strip()


# ─────────────────────────────────────────────────────────────────────────────
# Context extraction (handles all 3 result formats)
# ─────────────────────────────────────────────────────────────────────────────
def get_context(record: dict) -> str:
    if record.get("compressed"):
        chunks = record["compressed"]
    elif record.get("retrieved"):
        chunks = record["retrieved"]
    else:
        return ""
    return "\n\n".join(c.get("text", "") for c in chunks)


# ─────────────────────────────────────────────────────────────────────────────
# Evaluate one version
# ─────────────────────────────────────────────────────────────────────────────
def evaluate_version(name: str, path: str, answer_key: dict[int, str]) -> dict:
    print(f"\n{'=' * 60}")
    print(f"Scoring: {name}")
    print("=" * 60)

    with open(path, encoding="utf-8") as f:
        results = json.load(f)

    normal_scores = []
    trap_scores   = []
    skipped       = []

    for r in results:
        num = r["num"]
        gt  = answer_key.get(num, "")

        if not gt:
            skipped.append(num)
            print(f"  Q{num:<2} SKIPPED — no answer key")
            continue

        if is_trap(gt):
            s = score_trap(r["question"], r["answer"])
            trap_scores.append(s)
            print(f"  Q{num:<2} [trap] refused_correctly={s['_refused']:.0f}")
        else:
            s = score_normal(r["question"], r["answer"], get_context(r), gt)
            normal_scores.append(s)
            print(f"  Q{num:<2} faith={s['faithfulness']:.2f} "
                  f"rel={s['answer_relevancy']:.2f} "
                  f"prec={s['context_precision']:.2f} "
                  f"rec={s['context_recall']:.2f}")
        time.sleep(0.3)

    # averages over ANSWERABLE questions only
    def avg(key):
        if not normal_scores:
            return 0.0
        return sum(s[key] for s in normal_scores) / len(normal_scores)

    trap_pass = sum(s["_refused"] for s in trap_scores)
    return {
        "faithfulness":      round(avg("faithfulness"), 3),
        "answer_relevancy":  round(avg("answer_relevancy"), 3),
        "context_precision": round(avg("context_precision"), 3),
        "context_recall":    round(avg("context_recall"), 3),
        "answerable_count":  len(normal_scores),
        "trap_passed":       f"{int(trap_pass)}/{len(trap_scores)}" if trap_scores else "n/a",
        "skipped":           skipped,
    }


def main():
    print("Loading answer key...")
    answer_key = load_answer_key(EVAL_XLSX)
    filled = sum(1 for v in answer_key.values() if v)
    print(f"  {filled}/{len(answer_key)} answer keys filled")
    if filled < len(answer_key):
        print("  WARNING: empty answer keys will be SKIPPED, not guessed.")
        print("  Fill column E in the spreadsheet for complete results.")

    all_scores = {}
    for name, path in RESULT_FILES.items():
        try:
            all_scores[name] = evaluate_version(name, path, answer_key)
        except FileNotFoundError:
            print(f"  {path} not found — skipped")

    # ── Comparison table ──
    print("\n\n" + "=" * 70)
    print("FINAL COMPARISON — answerable questions only (traps reported separately)")
    print("=" * 70)
    metrics = ["faithfulness", "answer_relevancy", "context_precision", "context_recall"]
    print(f"{'Metric':<20}" + "".join(f"{n:>14}" for n in all_scores))
    print("-" * 70)
    for m in metrics:
        print(f"{m:<20}" + "".join(f"{all_scores[n][m]:>14.3f}" for n in all_scores))
    print("-" * 70)
    print(f"{'OVERALL':<20}" + "".join(
        f"{sum(all_scores[n][m] for m in metrics)/4:>14.3f}" for n in all_scores))
    print("=" * 70)
    print(f"{'Answerable Qs':<20}" + "".join(f"{all_scores[n]['answerable_count']:>14}" for n in all_scores))
    print(f"{'Traps refused':<20}" + "".join(f"{all_scores[n]['trap_passed']:>14}" for n in all_scores))

    with open("eval_report_v2.json", "w", encoding="utf-8") as f:
        json.dump(all_scores, f, indent=2)
    print("\nSaved to eval_report_v2.json")


if __name__ == "__main__":
    main()