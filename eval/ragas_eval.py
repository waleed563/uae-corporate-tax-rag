"""
ragas_eval.py — Compare Baseline vs Ring 1 vs Ring 2 using RAGAS metrics.

WHAT THIS SCRIPT DOES:
─────────────────────────────────────────────────────────────────────────
Loads the three results JSON files and runs RAGAS evaluation on each.
Produces a side-by-side comparison table showing which version won
on each metric, and saves a full report to ragas_report.md.

RAGAS METRICS EXPLAINED (simply):
─────────────────────────────────────────────────────────────────────────
1. FAITHFULNESS (0-1)
   "Is the answer actually supported by the retrieved chunks?"
   Score 1.0 = every claim in the answer comes from the context.
   Score 0.5 = half the claims are made up / hallucinated.
   Does NOT need an answer key. Uses LLM to check.

2. ANSWER RELEVANCY (0-1)
   "Does the answer actually address the question?"
   Score 1.0 = directly on-topic.
   Score 0.5 = partially answers it, goes off-topic, too vague.
   Does NOT need an answer key. Uses embeddings to check.

3. CONTEXT PRECISION (0-1)  ← needs answer key
   "Were the retrieved chunks actually useful?"
   Score 1.0 = every retrieved chunk was relevant.
   Score 0.5 = half the retrieved chunks were noise.
   Needs ground_truth to judge relevance.

4. CONTEXT RECALL (0-1)  ← needs answer key
   "Did retrieval find ALL the information needed to answer?"
   Score 1.0 = everything needed was in the retrieved chunks.
   Score 0.5 = half the necessary information was missing.
   Needs ground_truth to judge completeness.

ANSWER KEY:
─────────────────────────────────────────────────────────────────────────
If column E ("Answer key") in uae_tax_rag_eval.xlsx is filled,
this script runs all 4 metrics. If it's empty, it runs only
faithfulness + answer_relevancy (still meaningful).

Fill the answer key from the PDF for the most complete evaluation.

HOW TO RUN:
─────────────────────────────────────────────────────────────────────────
    python ragas_eval.py \\
        --baseline baseline_results.json \\
        --ring1    ring1_results.json \\
        --ring2    ring2_results.json \\
        --eval     uae_tax_rag_eval.xlsx

LEARNING EXERCISE — before running:
─────────────────────────────────────────────────────────────────────────
Predict the scores BEFORE you run. Based on what you know:

1. Which version do you expect to score highest on FAITHFULNESS?
   Hint: which version has the strictest guardrails against hallucination?

2. Which version do you expect to score highest on CONTEXT PRECISION?
   Hint: which version actively removes noisy sentences from chunks?

3. Why might ANSWER RELEVANCY be similar across all three versions
   even though the quality improved a lot?
   Hint: all three use the same LLM and same grounding prompt structure.

Write your predictions, then run.
"""

import argparse
import json
import os
import sys

import pandas as pd
from datasets import Dataset
from openpyxl import load_workbook
from ragas import evaluate
from ragas.metrics import (
    answer_relevancy,
    context_precision,
    context_recall,
    faithfulness,
)

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ─────────────────────────────────────────────────────────────────────────────
# Load results JSON
# ─────────────────────────────────────────────────────────────────────────────
def load_results(path: str) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# ─────────────────────────────────────────────────────────────────────────────
# Load answer key from xlsx (column E)
# ─────────────────────────────────────────────────────────────────────────────
def load_answer_key(xlsx_path: str) -> dict[int, str]:
    """
    Returns {question_num: answer_key_text}.
    Empty cells return None — we skip those questions for precision/recall.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Test Set"]
    keys = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, question, category, what_tested, answer_key = (
            row[0], row[1], row[2], row[3], row[4]
        )
        if num:
            keys[num] = answer_key  # None if cell is empty
    wb.close()
    return keys


# ─────────────────────────────────────────────────────────────────────────────
# Build RAGAS dataset from results
# ─────────────────────────────────────────────────────────────────────────────
def build_dataset(
    results: list[dict],
    answer_key: dict[int, str],
    use_ground_truth: bool,
) -> Dataset:
    """
    RAGAS expects columns:
      question   : str
      answer     : str
      contexts   : list[str]   — the retrieved chunks as plain text
      ground_truth: str        — only needed for precision/recall
    """
    questions    = []
    answers      = []
    contexts     = []
    ground_truths = []

    for r in results:
        questions.append(r["question"])
        answers.append(r["answer"])

        # Pull chunk texts — handle both baseline (retrieved) and ring1/2 (compressed)
        # Ring 2 has "compressed" key with the cleaned chunks
        # Ring 1 has "retrieved" key
        # Baseline has "retrieved" key
        if "compressed" in r and r["compressed"]:
            chunk_texts = [c["text"] for c in r["compressed"]]
        elif "retrieved" in r and r["retrieved"]:
            chunk_texts = [c["text"] for c in r["retrieved"]]
        else:
            chunk_texts = [""]

        contexts.append(chunk_texts)

        gt = answer_key.get(r["num"], None) if use_ground_truth else None
        ground_truths.append(gt or "")

    data = {
        "question":    questions,
        "answer":      answers,
        "contexts":    contexts,
        "ground_truth": ground_truths,
    }
    return Dataset.from_dict(data)


# ─────────────────────────────────────────────────────────────────────────────
# Run RAGAS on one dataset
# ─────────────────────────────────────────────────────────────────────────────
def run_ragas(dataset: Dataset, use_ground_truth: bool) -> dict:
    metrics = [faithfulness, answer_relevancy]
    if use_ground_truth:
        metrics += [context_precision, context_recall]

    result = evaluate(dataset, metrics=metrics)

    scores = {
        "faithfulness":    round(result["faithfulness"], 4),
        "answer_relevancy": round(result["answer_relevancy"], 4),
    }
    if use_ground_truth:
        scores["context_precision"] = round(result["context_precision"], 4)
        scores["context_recall"]    = round(result["context_recall"], 4)

    return scores


# ─────────────────────────────────────────────────────────────────────────────
# Write markdown report
# ─────────────────────────────────────────────────────────────────────────────
def write_report(
    scores: dict[str, dict],
    use_ground_truth: bool,
    path: str,
):
    lines = ["# RAGAS Evaluation Report — UAE Tax RAG", ""]
    lines.append("## Score comparison")
    lines.append("")

    metrics = ["faithfulness", "answer_relevancy"]
    if use_ground_truth:
        metrics += ["context_precision", "context_recall"]

    # Header
    lines.append("| Metric | Baseline | Ring 1 | Ring 2 | Winner |")
    lines.append("|---|---|---|---|---|")

    for m in metrics:
        b = scores["baseline"].get(m, "n/a")
        r1 = scores["ring1"].get(m, "n/a")
        r2 = scores["ring2"].get(m, "n/a")
        vals = {"baseline": b, "ring1": r1, "ring2": r2}
        winner = max(vals, key=lambda k: vals[k] if isinstance(vals[k], float) else -1)
        lines.append(f"| {m} | {b} | {r1} | {r2} | **{winner}** |")

    lines.append("")
    lines.append("## What each metric means")
    lines.append("")
    lines.append("**Faithfulness** — are the answers grounded in the retrieved context?")
    lines.append("Score of 1.0 = no hallucinations. Score below 0.9 = hallucinations present.")
    lines.append("")
    lines.append("**Answer relevancy** — does the answer actually address the question?")
    lines.append("Score of 1.0 = fully on-topic. Lower scores = vague or off-topic answers.")
    lines.append("")

    if use_ground_truth:
        lines.append("**Context precision** — were the retrieved chunks actually useful?")
        lines.append("Score of 1.0 = no noisy chunks. Lower = too much irrelevant context retrieved.")
        lines.append("")
        lines.append("**Context recall** — did retrieval find everything needed?")
        lines.append("Score of 1.0 = nothing important was missed. Lower = key info wasn't retrieved.")
        lines.append("")

    lines.append("## How to read this for your portfolio")
    lines.append("")
    lines.append("The story is the PROGRESSION across versions:")
    lines.append("- Faithfulness should go UP (Ring 2's Rule B prevents hallucination)")
    lines.append("- Answer relevancy should stay HIGH (all three use the same LLM)")
    lines.append("- Context precision should go UP (compression removes noisy sentences)")
    lines.append("- Context recall should go UP (multi-query retrieves more relevant chunks)")
    lines.append("")
    lines.append("If any metric goes DOWN from Ring 1 to Ring 2, that is honest data.")
    lines.append("Explain WHY in your LinkedIn post — that is more credible than cherry-picked wins.")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--baseline", required=True)
    ap.add_argument("--ring1",    required=True)
    ap.add_argument("--ring2",    required=True)
    ap.add_argument("--eval",     required=True, help="xlsx with answer key")
    args = ap.parse_args()

    if not os.getenv("OPENAI_API_KEY"):
        sys.exit("Set OPENAI_API_KEY first.")

    print("Loading answer key from spreadsheet...")
    answer_key = load_answer_key(args.eval)
    filled     = sum(1 for v in answer_key.values() if v)
    total      = len(answer_key)
    print(f"  {filled}/{total} questions have an answer key filled")

    use_ground_truth = filled >= 10  # need at least 10 to be meaningful
    if use_ground_truth:
        print("  Running all 4 metrics (faithfulness, relevancy, precision, recall)")
    else:
        print("  Running 2 metrics only (faithfulness, relevancy)")
        print("  Fill the answer key column in the xlsx for precision + recall")

    print("\nLoading results...")
    baseline_results = load_results(args.baseline)
    ring1_results    = load_results(args.ring1)
    ring2_results    = load_results(args.ring2)
    print(f"  Baseline: {len(baseline_results)} questions")
    print(f"  Ring 1:   {len(ring1_results)} questions")
    print(f"  Ring 2:   {len(ring2_results)} questions")

    print("\nBuilding RAGAS datasets...")
    ds_baseline = build_dataset(baseline_results, answer_key, use_ground_truth)
    ds_ring1    = build_dataset(ring1_results,    answer_key, use_ground_truth)
    ds_ring2    = build_dataset(ring2_results,    answer_key, use_ground_truth)

    print("\nRunning RAGAS on Baseline...")
    scores_baseline = run_ragas(ds_baseline, use_ground_truth)
    print(f"  {scores_baseline}")

    print("\nRunning RAGAS on Ring 1...")
    scores_ring1 = run_ragas(ds_ring1, use_ground_truth)
    print(f"  {scores_ring1}")

    print("\nRunning RAGAS on Ring 2...")
    scores_ring2 = run_ragas(ds_ring2, use_ground_truth)
    print(f"  {scores_ring2}")

    all_scores = {
        "baseline": scores_baseline,
        "ring1":    scores_ring1,
        "ring2":    scores_ring2,
    }

    print("\n" + "─" * 60)
    print("FINAL COMPARISON")
    print("─" * 60)
    metrics = ["faithfulness", "answer_relevancy"]
    if use_ground_truth:
        metrics += ["context_precision", "context_recall"]

    for m in metrics:
        b  = scores_baseline.get(m, "n/a")
        r1 = scores_ring1.get(m, "n/a")
        r2 = scores_ring2.get(m, "n/a")
        print(f"{m:25s}  Baseline: {b}  Ring1: {r1}  Ring2: {r2}")

    write_report(all_scores, use_ground_truth, "ragas_report.md")
    print("\nFull report saved to ragas_report.md")
    print("Bring that file back and we will read the results together.")


if __name__ == "__main__":
    main()